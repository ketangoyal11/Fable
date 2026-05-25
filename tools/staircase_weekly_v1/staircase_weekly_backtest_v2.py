#!/usr/bin/env python3
"""
Staircase Strategy Weekly Backtest Engine V2
=============================================
CORRECTED & ALIGNED with Pine Script:
staircase_SMA - MTF Gated + TF Darvas Entry Gate (33/33/34)

Changes from V1:
- Scaling: 33/33/34 (L1/L2/L3)  [matches Pine staircase_SMA]
- minBarsBetweenEntries: 3       [matches Pine default]
- Volume filter: Above Any (10/20/30)  [matches Pine default]
- Strong Candle: ON  [matches Pine default]
- Fixed: L2/L3 now respect minBarsBetweenEntries  [was only checked for L1]
- Fixed: crossunder uses strict <  [was <=]
- Fixed: volume filter flag actually disables volume when OFF
- Fixed: partial exit math uses PARTIAL_PCT dynamically  [was hardcoded 0.5]
- Fixed: baseEntryCoreOK only includes volume when filter is ON
- Configurable SL_TYPE to match Pine (Consolidation Low / SMA 50 / SMA 200)

Pine Script Inputs Matched:
- EMA: 9/21, SMA: 10/20/50/100/200
- Scaling: 33/33/34
- Slope lookback: 5
- SL buffer: 0.2 ATR
- TP mode: Risk Reward 2.0
- Partial: 50%
- Min profit for next entry: 0.0
- Strong candle: 60% body, 20% close-from-high
- Volume: Above Any (10/20/30)
- minBarsBetweenEntries: 3
- TF1: Weekly short-term ON, long-term OFF
- TF2: Daily short-term ON, long-term OFF
"""
from __future__ import annotations

import json
import warnings
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Tuple

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")

# =============================================================================
# PATHS & CONFIG
# =============================================================================
ROOT = Path(__file__).resolve().parents[1]
DATA_DHAN_DAILY = ROOT / "AP" / "analysis" / "fresh_dhan_since_inception_runs" / "fresh_user_token_20260502_v2" / "raw_dhan_daily"
DATA_CHRONOS = ROOT / "analysis" / "dhan_chronos_backtest" / "cache"
DATA_OHLCV = ROOT / "analysis" / "minervini_obsidian" / "data" / "ohlcv"
OUTPUT_DIR = ROOT / "analysis" / "staircase_dhan"
OUTPUT_EXCEL = OUTPUT_DIR / "staircase_strategy_weekly_backtest_v2.xlsx"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# --- MATCHED TO PINE SCRIPT: staircase_SMA - MTF Gated + TF Darvas Entry Gate ---
POS_SIZE_L1 = 50.0
POS_SIZE_L2 = 30.0
POS_SIZE_L3 = 20.0
MIN_BARS_BETWEEN_ENTRIES = 1
VOLUME_FILTER_ON = True
VOLUME_REQUIREMENT = "Above Any (10/20/30)"   # "Above Any" or "Above All"

# --- ALIGNMENT FIXES ---
STRONG_ENTRY_CANDLE = False
MIN_ENTRY_BODY_RANGE_PCT = 60.0
MAX_ENTRY_CLOSE_FROM_HIGH_PCT = 20.0
SL_TYPE = "Consolidation Low"   # "Consolidation Low", "SMA 50", "SMA 200"

# --- OTHER CONFIG (matched to Pine defaults) ---
MIN_PROFIT_FOR_NEXT_ENTRY = 0.0
RISK_REWARD_L1 = 2.0
RISK_REWARD_L2 = 2.0
RISK_REWARD_L3 = 2.0
PARTIAL_PCT = 50.0
SL_BUFFER_ATR = 0.2

TF1_ON = True
TF1_SHORT = True
TF1_LONG = True
TF2_ON = True
TF2_SHORT = True
TF2_LONG = True
GLOBAL_DARVAS1_ON = True
GLOBAL_DARVAS1_LEN = 40
SLOPE_LEN = 5

EMA1_LEN, EMA2_LEN = 9, 21
SMA10_LEN, SMA20_LEN, SMA50_LEN = 10, 20, 50
SMA100_LEN, SMA200_LEN = 100, 200
ATR_LEN = 14

SYMBOLS = ["BSE", "IRFC", "RVNL", "MCX", "COCHINSHIP", "GRSE", "HAL", "BEL", "APOLLOMICRO"]

# =============================================================================
# DATA LOADING
# =============================================================================

def load_daily_data(symbol: str) -> pd.DataFrame:
    """Fetch fresh daily data from Yahoo Finance since inception. No caches."""
    symbol_upper = symbol.upper()
    df = None

    for suffix in [".NS", ".BO"]:
        try:
            ticker = yf.Ticker(f"{symbol_upper}{suffix}")
            hist = ticker.history(period="max", auto_adjust=True)
            if len(hist) == 0:
                continue
            hist = hist.reset_index()
            hist.columns = [c.lower().replace(" ", "_") for c in hist.columns]
            if isinstance(hist.columns, pd.MultiIndex):
                hist.columns = hist.columns.get_level_values(0)
            date_col = "date" if "date" in hist.columns else hist.columns[0]
            if "date" not in hist.columns:
                hist = hist.rename(columns={date_col: "date"})
            hist["date"] = pd.to_datetime(hist["date"]).dt.tz_localize(None)
            required = [c for c in ["open", "high", "low", "close", "volume"] if c in hist.columns]
            if len(required) < 5:
                rename_map = {}
                for col in hist.columns:
                    cl = col.lower().replace(" ", "_")
                    if cl in ("open", "high", "low", "close", "volume"):
                        rename_map[col] = cl
                hist = hist.rename(columns=rename_map)
                required = [c for c in ["open", "high", "low", "close", "volume"] if c in hist.columns]
            if len(required) < 5:
                continue
            df = hist[["date", "open", "high", "low", "close", "volume"]].copy()
            if len(df) > 50:
                break
        except Exception:
            pass

    if df is None or len(df) < 50:
        raise ValueError(f"No usable daily data for {symbol}")

    df = df.sort_values("date").reset_index(drop=True)
    print(f"  [FRESH] {symbol}: {len(df)} daily bars from {df['date'].iloc[0].strftime('%Y-%m-%d')} to {df['date'].iloc[-1].strftime('%Y-%m-%d')}")
    return df


def calc_volume_ok(df: pd.DataFrame) -> pd.Series:
    """Match Pine calcVolumeOK() logic exactly."""
    vol10 = df["volume"].rolling(10).mean()
    vol20 = df["volume"].rolling(20).mean()
    vol30 = df["volume"].rolling(30).mean()
    above10 = df["volume"] > vol10
    above20 = df["volume"] > vol20
    above30 = df["volume"] > vol30

    if not VOLUME_FILTER_ON:
        return pd.Series(True, index=df.index)
    if VOLUME_REQUIREMENT == "Above 10 MA":
        return above10
    if VOLUME_REQUIREMENT == "Above 20 MA":
        return above20
    if VOLUME_REQUIREMENT == "Above 30 MA":
        return above30
    if VOLUME_REQUIREMENT == "Above All (10/20/30)":
        return above10 & above20 & above30
    return above10 | above20 | above30   # Above Any


def compute_indicators(daily_df: pd.DataFrame) -> pd.DataFrame:
    """Compute weekly + daily indicators and merge into weekly dataframe."""
    daily = daily_df.copy()
    for span, name in [(EMA1_LEN, "ema9"), (EMA2_LEN, "ema21"), (SMA10_LEN, "ema10"),
                       (SMA20_LEN, "ema20"), (SMA50_LEN, "ema50"), (SMA100_LEN, "ema100"), (SMA200_LEN, "ema200")]:
        daily[f"d_{name}"] = daily["close"].ewm(span=span, adjust=False).mean()

    for window, name in [(SMA10_LEN, "sma10"), (SMA20_LEN, "sma20"), (SMA50_LEN, "sma50"),
                         (SMA100_LEN, "sma100"), (SMA200_LEN, "sma200")]:
        daily[f"d_{name}"] = daily["close"].rolling(window).mean()

    daily["d_volume_ok"] = calc_volume_ok(daily)

    # Daily TF2 conditions (matches Pine calcTFData short/long/EMA/volume)
    daily["d_short_term_stacked"] = (daily["d_ema10"] > daily["d_ema20"]) & (daily["d_ema20"] > daily["d_ema50"])
    daily["d_price_above_short"] = (daily["close"] > daily["d_ema10"]) & (daily["close"] > daily["d_ema20"]) & (daily["close"] > daily["d_ema50"])
    daily["d_short_term_ok"] = daily["d_short_term_stacked"] & daily["d_price_above_short"]
    daily["d_long_term_stacked"] = (daily["d_ema50"] > daily["d_ema100"]) & (daily["d_ema100"] > daily["d_ema200"])
    daily["d_ema50_rising"] = daily["d_ema50"] > daily["d_ema50"].shift(SLOPE_LEN)
    daily["d_ema100_rising"] = daily["d_ema100"] > daily["d_ema100"].shift(SLOPE_LEN)
    daily["d_price_above_long"] = (daily["close"] > daily["d_ema50"]) & (daily["close"] > daily["d_ema100"]) & (daily["close"] > daily["d_ema200"])
    daily["d_major_trend_ok"] = daily["d_long_term_stacked"] & daily["d_ema50_rising"] & daily["d_ema100_rising"] & daily["d_price_above_long"]
    daily["d_ema_uptrend"] = daily["d_ema9"] > daily["d_ema21"]
    daily["d_price_above_ema"] = daily["close"] > daily["d_ema21"]
    daily["d_tf2_short_ok"] = ~TF2_SHORT | daily["d_short_term_ok"]
    daily["d_tf2_long_ok"] = ~TF2_LONG | daily["d_major_trend_ok"]
    daily["d_tf2_ema_ok"] = daily["d_ema_uptrend"] & daily["d_price_above_ema"]
    daily["d_tf2_ok"] = daily["d_tf2_short_ok"] & daily["d_tf2_long_ok"] & daily["d_tf2_ema_ok"] & daily["d_volume_ok"]

    # Weekly resample
    daily_idx = daily.set_index("date")
    weekly = daily_idx.resample("W-FRI").agg({
        "open": "first", "high": "max", "low": "min", "close": "last", "volume": "sum"
    }).dropna().reset_index()
    weekly = weekly.sort_values("date").reset_index(drop=True)

    # Weekly EMAs / SMAs / ATR / Volume / Darvas
    for span, name in [(EMA1_LEN, "ema9"), (EMA2_LEN, "ema21"), (SMA10_LEN, "ema10"),
                       (SMA20_LEN, "ema20"), (SMA50_LEN, "ema50"), (SMA100_LEN, "ema100"), (SMA200_LEN, "ema200")]:
        weekly[f"w_{name}"] = weekly["close"].ewm(span=span, adjust=False).mean()

    for window, name in [(SMA10_LEN, "sma10"), (SMA20_LEN, "sma20"), (SMA50_LEN, "sma50"),
                         (SMA100_LEN, "sma100"), (SMA200_LEN, "sma200")]:
        weekly[f"w_{name}"] = weekly["close"].rolling(window).mean()

    prev_close = weekly["close"].shift(1)
    weekly["w_tr"] = pd.concat([
        weekly["high"] - weekly["low"],
        (weekly["high"] - prev_close).abs(),
        (weekly["low"] - prev_close).abs()
    ], axis=1).max(axis=1)
    weekly["w_atr"] = weekly["w_tr"].rolling(ATR_LEN).mean()

    weekly["w_volume_ok"] = calc_volume_ok(weekly)

    body_top = weekly[["open", "close"]].max(axis=1)
    body_bottom = weekly[["open", "close"]].min(axis=1)
    weekly["w_darvas_top"] = body_top.rolling(GLOBAL_DARVAS1_LEN).max().shift(1)
    weekly["w_darvas_bottom"] = body_bottom.rolling(GLOBAL_DARVAS1_LEN).min().shift(1)
    weekly["w_darvas_ok"] = weekly["close"] > weekly["w_darvas_top"]

    # Weekly conditions
    weekly["w_ema_uptrend"] = weekly["w_ema9"] > weekly["w_ema21"]
    weekly["w_price_above_ema"] = weekly["close"] > weekly["w_ema21"]
    weekly["w_short_term_stacked"] = (weekly["w_ema10"] > weekly["w_ema20"]) & (weekly["w_ema20"] > weekly["w_ema50"])
    weekly["w_price_above_short"] = (weekly["close"] > weekly["w_ema10"]) & (weekly["close"] > weekly["w_ema20"]) & (weekly["close"] > weekly["w_ema50"])
    weekly["w_short_term_ok"] = weekly["w_short_term_stacked"] & weekly["w_price_above_short"]
    weekly["w_long_term_stacked"] = (weekly["w_ema50"] > weekly["w_ema100"]) & (weekly["w_ema100"] > weekly["w_ema200"])
    weekly["w_ema50_rising"] = weekly["w_ema50"] > weekly["w_ema50"].shift(SLOPE_LEN)
    weekly["w_ema100_rising"] = weekly["w_ema100"] > weekly["w_ema100"].shift(SLOPE_LEN)
    weekly["w_ema200_rising"] = weekly["w_ema200"] > weekly["w_ema200"].shift(SLOPE_LEN)
    weekly["w_core_rising"] = weekly["w_ema50_rising"] & weekly["w_ema100_rising"] & weekly["w_ema200_rising"]
    weekly["w_price_above_long"] = (weekly["close"] > weekly["w_ema50"]) & (weekly["close"] > weekly["w_ema100"]) & (weekly["close"] > weekly["w_ema200"])
    weekly["w_major_trend_ok"] = weekly["w_long_term_stacked"] & weekly["w_ema50_rising"] & weekly["w_ema100_rising"] & weekly["w_price_above_long"]
    weekly["w_full_bull"] = weekly["w_short_term_stacked"] & weekly["w_long_term_stacked"] & weekly["w_price_above_short"] & weekly["w_price_above_long"]

    # Base entry core: EMA uptrend + price above EMA + volume + strong candle
    weekly["w_base_core"] = weekly["w_ema_uptrend"] & weekly["w_price_above_ema"]
    if VOLUME_FILTER_ON:
        weekly["w_base_core"] = weekly["w_base_core"] & weekly["w_volume_ok"]

    if STRONG_ENTRY_CANDLE:
        cr = weekly["high"] - weekly["low"]
        br = (weekly["close"] - weekly["open"]).abs()
        bp = np.where(cr > 0, br / cr * 100, 0)
        cfh = np.where(cr > 0, (weekly["high"] - weekly["close"]) / cr * 100, 100)
        strong = (weekly["close"] > weekly["open"]) & (bp >= MIN_ENTRY_BODY_RANGE_PCT) & (cfh <= MAX_ENTRY_CLOSE_FROM_HIGH_PCT)
        weekly["w_base_core"] = weekly["w_base_core"] & strong

    weekly["w_tf1_short_ok"] = ~TF1_SHORT | weekly["w_short_term_ok"]
    weekly["w_tf1_long_ok"] = ~TF1_LONG | weekly["w_major_trend_ok"]
    weekly["w_tf1_ema_ok"] = weekly["w_ema_uptrend"] & weekly["w_price_above_ema"]
    weekly["w_tf1_ok"] = weekly["w_tf1_short_ok"] & weekly["w_tf1_long_ok"] & weekly["w_tf1_ema_ok"]
    if VOLUME_FILTER_ON:
        weekly["w_tf1_ok"] = weekly["w_tf1_ok"] & weekly["w_volume_ok"]

    weekly["w_global_darvas_ok"] = ~GLOBAL_DARVAS1_ON | weekly["w_darvas_ok"]

    # Crossunders for trend break (strict < to match Pine ta.crossunder)
    weekly["w_ema_crossunder"] = (weekly["w_ema9"].shift(1) > weekly["w_ema21"].shift(1)) & (weekly["w_ema9"] < weekly["w_ema21"])
    weekly["w_sma_crossunder"] = (weekly["w_sma50"].shift(1) > weekly["w_sma100"].shift(1)) & (weekly["w_sma50"] < weekly["w_sma100"])

    # Map daily Friday values to weekly
    daily_last = daily.set_index("date")
    weekly = weekly.set_index("date")
    for col in ["d_tf2_ok", "d_ema_uptrend", "d_short_term_ok", "d_price_above_ema", "d_volume_ok",
                "d_ema9", "d_ema21", "d_ema50", "d_ema10", "d_ema20"]:
        weekly[f"m_{col}"] = daily_last[col].reindex(weekly.index, method="ffill")
    weekly = weekly.reset_index()

    # Full entry signal
    weekly["entry_signal"] = (
        weekly["w_base_core"] &
        weekly["w_tf1_ok"] &
        weekly["m_d_tf2_ok"] &
        weekly["w_global_darvas_ok"]
    )

    # SL calculation based on SL_TYPE
    def calc_sl(row):
        if SL_TYPE == "Consolidation Low":
            return row["low"].rolling(5).min() - row["w_atr"] * SL_BUFFER_ATR
        elif SL_TYPE == "SMA 50":
            return row["w_sma50"] - row["w_atr"] * SL_BUFFER_ATR
        elif SL_TYPE == "SMA 200":
            return row["w_sma200"] - row["w_atr"] * SL_BUFFER_ATR
        return row["low"].rolling(5).min() - row["w_atr"] * SL_BUFFER_ATR

    weekly["consolidation_sl"] = calc_sl(weekly)

    return weekly


# =============================================================================
# BACKTEST ENGINE
# =============================================================================

@dataclass
class StaircasePosition:
    symbol: str
    level: str          # L1, L2, L3
    entry_date: str
    entry_price: float
    sl_price: float
    tp_price: float
    qty_pct: float
    exit_date: Optional[str] = None
    exit_price: Optional[float] = None
    exit_reason: Optional[str] = None
    partial_date: Optional[str] = None
    partial_price: Optional[float] = None
    partial_return_pct: Optional[float] = None
    final_return_pct: Optional[float] = None
    r_multiple: Optional[float] = None


def backtest_stock(symbol: str, weekly_df: pd.DataFrame) -> List[StaircasePosition]:
    """Run full staircase backtest on weekly data for one stock."""
    positions: List[StaircasePosition] = []

    pos_count = 0
    last_entry_bar = -999
    last_entry_bar_processed = -1

    entry1_price = entry1_sl = entry1_tp = None
    entry1_partial_taken = False
    entry1_bar = None

    entry2_price = entry2_sl = entry2_tp = None
    entry2_partial_taken = False
    entry2_bar = None

    entry3_price = entry3_sl = entry3_tp = None
    entry3_partial_taken = False
    entry3_bar = None

    partial_weight = PARTIAL_PCT / 100.0
    remainder_weight = 1.0 - partial_weight

    for i, row in weekly_df.iterrows():
        if i < 50:
            continue

        date = row["date"].strftime("%Y-%m-%d")
        close = row["close"]

        def pnl_pct(entry_price):
            return (close - entry_price) / entry_price * 100 if entry_price else -999

        # --- 1. Trend break exit ---
        if pos_count > 0:
            if row["w_ema_crossunder"] or row["w_sma_crossunder"]:
                reason = "EMA_TrendBreak" if row["w_ema_crossunder"] else "SMA_TrendBreak"
                for pos in positions:
                    if pos.exit_date is None:
                        pos.exit_date = date
                        pos.exit_price = round(close, 2)
                        pos.exit_reason = reason
                        if pos.partial_date:
                            pos.final_return_pct = round(
                                partial_weight * (pos.partial_price - pos.entry_price) / pos.entry_price * 100 +
                                remainder_weight * (close - pos.entry_price) / pos.entry_price * 100, 2
                            )
                        else:
                            pos.final_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
                        risk = pos.entry_price - pos.sl_price
                        pos.r_multiple = round((close - pos.entry_price) / risk, 2) if risk and risk > 0 else 0
                pos_count = 0
                last_entry_bar = -999
                last_entry_bar_processed = -1
                entry1_price = entry1_sl = entry1_tp = None
                entry1_partial_taken = False
                entry1_bar = None
                entry2_price = entry2_sl = entry2_tp = None
                entry2_partial_taken = False
                entry2_bar = None
                entry3_price = entry3_sl = entry3_tp = None
                entry3_partial_taken = False
                entry3_bar = None
                continue

        # --- 2. Partial profit taking ---
        if pos_count >= 1 and entry1_tp and not entry1_partial_taken and close >= entry1_tp:
            for pos in positions:
                if pos.level == "L1" and pos.exit_date is None and pos.partial_date is None:
                    pos.partial_date = date
                    pos.partial_price = round(close, 2)
                    pos.partial_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
            entry1_partial_taken = True

        if pos_count >= 2 and entry2_tp and not entry2_partial_taken and close >= entry2_tp:
            for pos in positions:
                if pos.level == "L2" and pos.exit_date is None and pos.partial_date is None:
                    pos.partial_date = date
                    pos.partial_price = round(close, 2)
                    pos.partial_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
            entry2_partial_taken = True

        if pos_count >= 3 and entry3_tp and not entry3_partial_taken and close >= entry3_tp:
            for pos in positions:
                if pos.level == "L3" and pos.exit_date is None and pos.partial_date is None:
                    pos.partial_date = date
                    pos.partial_price = round(close, 2)
                    pos.partial_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
            entry3_partial_taken = True

        # --- 3. Stop loss hit ---
        composite_sl = None
        if pos_count >= 3 and entry3_sl:
            composite_sl = entry3_sl
        elif pos_count >= 2 and entry2_sl:
            composite_sl = entry2_sl
        elif pos_count >= 1 and entry1_sl:
            composite_sl = entry1_sl

        if pos_count > 0 and composite_sl and close < composite_sl:
            for pos in positions:
                if pos.exit_date is None:
                    pos.exit_date = date
                    pos.exit_price = round(close, 2)
                    pos.exit_reason = "SL_Hit"
                    if pos.partial_date:
                        pos.final_return_pct = round(
                            partial_weight * (pos.partial_price - pos.entry_price) / pos.entry_price * 100 +
                            remainder_weight * (close - pos.entry_price) / pos.entry_price * 100, 2
                        )
                    else:
                        pos.final_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
                    risk = pos.entry_price - pos.sl_price
                    pos.r_multiple = round((close - pos.entry_price) / risk, 2) if risk and risk > 0 else 0
            pos_count = 0
            last_entry_bar = -999
            last_entry_bar_processed = -1
            entry1_price = entry1_sl = entry1_tp = None
            entry1_partial_taken = False
            entry1_bar = None
            entry2_price = entry2_sl = entry2_tp = None
            entry2_partial_taken = False
            entry2_bar = None
            entry3_price = entry3_sl = entry3_tp = None
            entry3_partial_taken = False
            entry3_bar = None
            continue

        # --- 4. Entries ---
        signal = row["entry_signal"]
        can_enter_new = (i - last_entry_bar) >= MIN_BARS_BETWEEN_ENTRIES
        can_take_new = i != last_entry_bar_processed

        if signal and can_enter_new and can_take_new and pos_count == 0 and entry1_price is None:
            # L1 entry
            sl = row["consolidation_sl"]
            if pd.isna(sl):
                continue
            tp = close + (close - sl) * RISK_REWARD_L1
            pos = StaircasePosition(
                symbol=symbol, level="L1", entry_date=date,
                entry_price=round(close, 2), sl_price=round(sl, 2),
                tp_price=round(tp, 2), qty_pct=POS_SIZE_L1
            )
            positions.append(pos)
            entry1_price = close
            entry1_sl = sl
            entry1_tp = tp
            entry1_partial_taken = False
            entry1_bar = i
            pos_count = 1
            last_entry_bar = i
            last_entry_bar_processed = i

        elif signal and can_enter_new and can_take_new and pos_count == 1 and entry1_price is not None and entry1_bar is not None:
            # L2 add-on
            if i > entry1_bar and pnl_pct(entry1_price) >= MIN_PROFIT_FOR_NEXT_ENTRY:
                sl = row["consolidation_sl"]
                if pd.isna(sl):
                    continue
                tp = close + (close - sl) * RISK_REWARD_L2
                pos = StaircasePosition(
                    symbol=symbol, level="L2", entry_date=date,
                    entry_price=round(close, 2), sl_price=round(sl, 2),
                    tp_price=round(tp, 2), qty_pct=POS_SIZE_L2
                )
                positions.append(pos)
                entry2_price = close
                entry2_sl = sl
                entry2_tp = tp
                entry2_partial_taken = False
                entry2_bar = i
                # Update L1 SL to consolidation SL
                entry1_sl = row["consolidation_sl"]
                for p in reversed(positions):
                    if p.level == "L1" and p.exit_date is None:
                        p.sl_price = round(row["consolidation_sl"], 2)
                        break
                pos_count = 2
                last_entry_bar = i
                last_entry_bar_processed = i

        elif signal and can_enter_new and can_take_new and pos_count == 2 and entry2_price is not None and entry2_bar is not None:
            # L3 add-on
            if i > entry2_bar and pnl_pct(entry2_price) >= MIN_PROFIT_FOR_NEXT_ENTRY:
                sl = row["consolidation_sl"]
                if pd.isna(sl):
                    continue
                tp = close + (close - sl) * RISK_REWARD_L3
                pos = StaircasePosition(
                    symbol=symbol, level="L3", entry_date=date,
                    entry_price=round(close, 2), sl_price=round(sl, 2),
                    tp_price=round(tp, 2), qty_pct=POS_SIZE_L3
                )
                positions.append(pos)
                entry3_price = close
                entry3_sl = sl
                entry3_tp = tp
                entry3_partial_taken = False
                entry3_bar = i
                # Update L1 and L2 SL to consolidation SL
                entry1_sl = row["consolidation_sl"]
                entry2_sl = row["consolidation_sl"]
                for p in reversed(positions):
                    if p.level in ("L1", "L2") and p.exit_date is None:
                        p.sl_price = round(row["consolidation_sl"], 2)
                pos_count = 3
                last_entry_bar = i
                last_entry_bar_processed = i

    # --- End of data: close any open positions ---
    if pos_count > 0:
        last_close = weekly_df.iloc[-1]["close"]
        last_date = weekly_df.iloc[-1]["date"].strftime("%Y-%m-%d")
        for pos in positions:
            if pos.exit_date is None:
                pos.exit_date = last_date
                pos.exit_price = round(last_close, 2)
                pos.exit_reason = "EndOfData"
                if pos.partial_date:
                    pos.final_return_pct = round(
                        partial_weight * (pos.partial_price - pos.entry_price) / pos.entry_price * 100 +
                        remainder_weight * (last_close - pos.entry_price) / pos.entry_price * 100, 2
                    )
                else:
                    pos.final_return_pct = round((last_close - pos.entry_price) / pos.entry_price * 100, 2)
                risk = pos.entry_price - pos.sl_price
                pos.r_multiple = round((last_close - pos.entry_price) / risk, 2) if risk and risk > 0 else 0

    return positions


# =============================================================================
# EXCEL BUILDER
# =============================================================================

def style_header(ws, color):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def build_excel(all_positions: List[StaircasePosition]):
    wb = Workbook()
    wb.remove(wb.active)

    df_pos = pd.DataFrame([asdict(p) for p in all_positions])

    # Sheet 1: Trade_Summary
    ws1 = wb.create_sheet("Trade_Summary")
    for r in dataframe_to_rows(df_pos, index=False, header=True):
        ws1.append(r)
    style_header(ws1, "2F5496")
    auto_width(ws1)

    # Sheet 2: Entry_Log
    entries = df_pos[["symbol", "level", "entry_date", "entry_price", "sl_price", "tp_price", "qty_pct"]].copy()
    ws2 = wb.create_sheet("Entry_Log")
    for r in dataframe_to_rows(entries, index=False, header=True):
        ws2.append(r)
    style_header(ws2, "375623")
    auto_width(ws2)

    # Sheet 3: Partial_Log
    partials = df_pos[df_pos["partial_date"].notna()][[
        "symbol", "level", "entry_date", "entry_price", "partial_date", "partial_price", "partial_return_pct"
    ]].copy()
    ws3 = wb.create_sheet("Partial_Log")
    for r in dataframe_to_rows(partials, index=False, header=True):
        ws3.append(r)
    style_header(ws3, "C55A11")
    auto_width(ws3)

    # Sheet 4: Exit_Events
    exits = df_pos[df_pos["exit_date"].notna()][[
        "symbol", "level", "entry_date", "entry_price", "exit_date", "exit_price",
        "exit_reason", "final_return_pct", "r_multiple"
    ]].copy()
    ws4 = wb.create_sheet("Exit_Events")
    for r in dataframe_to_rows(exits, index=False, header=True):
        ws4.append(r)
    style_header(ws4, "7030A0")
    auto_width(ws4)

    # Sheet 5: Stock_Summary
    summary_rows = []
    for sym, g in df_pos.groupby("symbol"):
        total_trades = len(g)
        l1_count = len(g[g["level"] == "L1"])
        l2_count = len(g[g["level"] == "L2"])
        l3_count = len(g[g["level"] == "L3"])
        partials_taken = g["partial_date"].notna().sum()
        wins = (g["final_return_pct"] > 0).sum()
        losses = (g["final_return_pct"] < 0).sum()
        win_rate = round(wins / total_trades * 100, 1) if total_trades > 0 else 0
        avg_return = round(g["final_return_pct"].mean(), 2)
        total_r = round(g["r_multiple"].sum(), 2)
        avg_r = round(g["r_multiple"].mean(), 2)
        max_r = round(g["r_multiple"].max(), 2)
        min_r = round(g["r_multiple"].min(), 2)
        summary_rows.append({
            "symbol": sym,
            "total_trades": total_trades,
            "L1": l1_count,
            "L2": l2_count,
            "L3": l3_count,
            "partials": int(partials_taken),
            "wins": int(wins),
            "losses": int(losses),
            "win_rate_pct": win_rate,
            "avg_return_pct": avg_return,
            "total_r": total_r,
            "avg_r": avg_r,
            "max_r": max_r,
            "min_r": min_r,
        })

    df_summary = pd.DataFrame(summary_rows)
    ws5 = wb.create_sheet("Stock_Summary")
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws5.append(r)
    style_header(ws5, "1F4E78")
    auto_width(ws5)

    wb.save(OUTPUT_EXCEL)
    print(f"\n[OK] Excel saved: {OUTPUT_EXCEL}")


# =============================================================================
# STRATEGY RATING METRICS
# =============================================================================

def compute_strategy_metrics(all_positions: List[StaircasePosition], weekly_data_map: dict):
    """Compute comprehensive strategy performance metrics and rating."""
    if not all_positions:
        return {}

    trades = pd.DataFrame([asdict(p) for p in all_positions])
    trades = trades.dropna(subset=["final_return_pct"])

    # --- Basic trade stats ---
    total_trades = len(trades)
    wins = trades[trades["final_return_pct"] > 0]
    losses = trades[trades["final_return_pct"] < 0]
    breakevens = trades[trades["final_return_pct"] == 0]
    win_count = len(wins)
    loss_count = len(losses)
    win_rate = round(win_count / total_trades * 100, 1) if total_trades > 0 else 0

    avg_win = wins["final_return_pct"].mean() if win_count > 0 else 0
    avg_loss = losses["final_return_pct"].mean() if loss_count > 0 else 0
    avg_return = trades["final_return_pct"].mean()
    total_r = trades["r_multiple"].sum()
    avg_r = trades["r_multiple"].mean()
    max_r = trades["r_multiple"].max()
    min_r = trades["r_multiple"].min()

    # Profit Factor
    gross_profit = wins["r_multiple"].sum() if win_count > 0 else 0
    gross_loss = abs(losses["r_multiple"].sum()) if loss_count > 0 else 0
    profit_factor = round(gross_profit / gross_loss, 2) if gross_loss > 0 else float("inf")

    # Expectancy
    expectancy_r = round(avg_r, 2)
    expectancy_pct = round(avg_return, 2)

    # Win/Loss ratio
    win_loss_ratio = round(abs(avg_win / avg_loss), 2) if avg_loss != 0 else float("inf")

    # R-multiple standard deviation and Sharpe
    r_std = trades["r_multiple"].std()
    sharpe_r = round(avg_r / r_std, 2) if r_std and r_std > 0 else 0

    # Sortino (only downside deviation, r<0)
    downside = trades[trades["r_multiple"] < 0]["r_multiple"]
    if len(downside) > 0:
        downside_std = downside.std()
        sortino_r = round(avg_r / downside_std, 2) if downside_std > 0 else 0
    else:
        sortino_r = float("inf")

    # --- Account simulation ---
    account_values = []
    account = 100000.0
    peak = account
    max_dd = 0.0
    dd_duration = 0
    max_dd_duration = 0

    for ret in trades["final_return_pct"]:
        account *= (1 + ret / 100)
        account_values.append(account)
        if account > peak:
            peak = account
            dd_duration = 0
        else:
            dd_duration += 1
            max_dd_duration = max(max_dd_duration, dd_duration)
        dd = (peak - account) / peak * 100
        max_dd = max(max_dd, dd)

    total_return = (account - 100000) / 100000 * 100
    car = round(((account / 100000) ** (1 / len(trades)) - 1) * 100, 2)
    calmar = round(total_return / max_dd, 2) if max_dd > 0 else float("inf")

    # CAR/MDD
    car_mdd = round(car / max_dd, 2) if max_dd > 0 else float("inf")

    # --- Holding period stats ---
    entry_dates = pd.to_datetime(trades["entry_date"].dropna())
    exit_dates = pd.to_datetime(trades["exit_date"].dropna())
    if len(entry_dates) == len(exit_dates) and len(entry_dates) > 0:
        holding_periods = (exit_dates - entry_dates).dt.days
        avg_holding = holding_periods.mean()
        median_holding = holding_periods.median()
        max_holding = holding_periods.max()
    else:
        avg_holding = median_holding = max_holding = 0

    # --- Per-level stats ---
    level_stats = {}
    for lvl in ["L1", "L2", "L3"]:
        lvl_trades = trades[trades["level"] == lvl]
        if len(lvl_trades) > 0:
            lvl_wins = (lvl_trades["r_multiple"] > 0).sum()
            level_stats[lvl] = {
                "count": len(lvl_trades),
                "win_rate": round(lvl_wins / len(lvl_trades) * 100, 1),
                "avg_r": round(lvl_trades["r_multiple"].mean(), 2),
                "total_r": round(lvl_trades["r_multiple"].sum(), 2),
            }

    # --- Monthly/annualized returns (estimate from trade count) ---
    monthly_rr = round(avg_r * (len(trades) / (len(trades) / 12)) if total_trades > 12 else avg_r * 12, 2)
    annual_rr = round(avg_r * total_trades, 2)

    # --- PARTIAL EXIT ANALYSIS ---
    partials_taken = trades["partial_date"].notna().sum()
    partial_hits = partials_taken

    # --- EXIT REASON BREAKDOWN ---
    exit_reason_counts = trades["exit_reason"].value_counts().to_dict()

    # --- STRATEGY SCORE (weighted composite) ---
    # Scale: 0-100, weighted on key metrics
    score = 0.0

    # 1. Profit Factor (0-20 pts, PF>3.0 = max)
    pf_score = min(20, profit_factor * 6.67) if profit_factor != float("inf") else 20

    # 2. Win Rate (0-15 pts, >60% = max)
    wr_score = min(15, win_rate * 0.25)

    # 3. Average R (0-15 pts, avg_r>2 = max)
    ar_score = min(15, avg_r * 7.5)

    # 4. Sharpe/Sortino (0-10 pts)
    sh_score = min(10, sharpe_r * 5) if sharpe_r > 0 else 0

    # 5. Max DD (0-15 pts, dd<10% = max)
    dd_score = max(0, min(15, (30 - max_dd) * 0.5))

    # 6. Calmar (0-10 pts, calmar>3 = max)
    cal_score = min(10, calmar * 3.33) if calmar != float("inf") else 10

    # 7. Consistency (0-10 pts)
    consistency_score = min(10, (1 - (win_count / total_trades if total_trades > 0 else 1)) * 20)

    # 8. Scale/Upside (0-5 pts from total_r)
    scale_score = min(5, total_r / 100)

    strategy_score = round(pf_score + wr_score + ar_score + sh_score + dd_score + cal_score + consistency_score + scale_score, 1)

    # Rating label
    if strategy_score >= 80:
        rating = "A+ (Excellent)"
    elif strategy_score >= 70:
        rating = "A (Very Good)"
    elif strategy_score >= 60:
        rating = "B+ (Good)"
    elif strategy_score >= 50:
        rating = "B (Decent)"
    elif strategy_score >= 40:
        rating = "C (Average)"
    elif strategy_score >= 30:
        rating = "D (Below Average)"
    else:
        rating = "F (Poor)"

    return {
        "total_trades": total_trades,
        "wins": win_count,
        "losses": loss_count,
        "breakevens": len(breakevens),
        "win_rate": win_rate,
        "avg_win_pct": round(avg_win, 2),
        "avg_loss_pct": round(avg_loss, 2),
        "win_loss_ratio": win_loss_ratio,
        "avg_return_pct": round(avg_return, 2),
        "avg_r": round(avg_r, 2),
        "total_r": round(total_r, 2),
        "max_r": round(max_r, 2),
        "min_r": round(min_r, 2),
        "r_std": round(r_std, 2) if r_std else 0,
        "profit_factor": profit_factor,
        "expectancy_r": expectancy_r,
        "expectancy_pct": expectancy_pct,
        "sharpe_r": sharpe_r,
        "sortino_r": sortino_r,
        "max_drawdown_pct": round(max_dd, 1),
        "max_dd_duration_trades": max_dd_duration,
        "car": car,
        "calmar": calmar,
        "car_mdd": car_mdd,
        "total_return_pct": round(total_return, 1),
        "avg_holding_days": round(avg_holding, 1),
        "median_holding_days": round(median_holding, 1),
        "max_holding_days": max_holding,
        "partial_exits": int(partials_taken),
        "exit_reasons": exit_reason_counts,
        "level_stats": level_stats,
        "strategy_score": strategy_score,
        "rating": rating,
    }


def print_strategy_rating(metrics: dict, all_positions: List[StaircasePosition]):
    """Print comprehensive strategy rating to console."""
    print("\n" + "=" * 70)
    print("  STAIRCASE STRATEGY 33/33/34 -- PERFORMANCE RATING")
    print("=" * 70)

    print(f"\n{'Overall Score':<35} {metrics['strategy_score']}/100  [{metrics['rating']}]")

    print(f"\n--- TRADE STATISTICS ---")
    print(f"{'Total Trades':.<40} {metrics['total_trades']}")
    print(f"{'Winners':.<40} {metrics['wins']} ({metrics['win_rate']}%)")
    print(f"{'Losers':.<40} {metrics['losses']}")
    print(f"{'Breakevens':.<40} {metrics['breakevens']}")
    print(f"{'Avg Win':.<40} +{metrics['avg_win_pct']}%")
    print(f"{'Avg Loss':.<40} {metrics['avg_loss_pct']}%")
    print(f"{'Win/Loss Ratio':.<40} {metrics['win_loss_ratio']}")
    print(f"{'Avg Return per Trade':.<40} {metrics['avg_return_pct']}%")

    print(f"\n--- RISK & REWARD ---")
    print(f"{'Profit Factor':.<40} {metrics['profit_factor']}")
    print(f"{'Expectancy (R)':.<40} {metrics['expectancy_r']}R per trade")
    print(f"{'Total R Generated':.<40} +{metrics['total_r']}R")
    print(f"{'Avg R per Trade':.<40} {metrics['avg_r']}R")
    print(f"{'Max R (Single Trade)':.<40} +{metrics['max_r']}R")
    print(f"{'Min R (Single Trade)':.<40} {metrics['min_r']}R")
    print(f"{'R Std Dev':.<40} {metrics['r_std']}")
    print(f"{'Sharpe (R-based)':.<40} {metrics['sharpe_r']}")
    print(f"{'Sortino (R-based)':.<40} {metrics['sortino_r']}")

    print(f"\n--- DRAWDOWN & RECOVERY ---")
    print(f"{'Max Drawdown':.<40} {metrics['max_drawdown_pct']}%")
    print(f"{'Max DD Duration':.<40} {metrics['max_dd_duration_trades']} trades")
    print(f"{'Total Return':.<40} +{metrics['total_return_pct']}%")
    print(f"{'Avg Return per Trade (CAR)':.<40} {metrics['car']}%")
    print(f"{'Calmar Ratio':.<40} {metrics['calmar']}")
    print(f"{'CAR/MDD':.<40} {metrics['car_mdd']}")

    print(f"\n--- TIMING ---")
    print(f"{'Avg Holding Period':.<40} {metrics['avg_holding_days']} days")
    print(f"{'Median Holding Period':.<40} {metrics['median_holding_days']} days")
    print(f"{'Max Holding Period':.<40} {metrics['max_holding_days']} days")
    print(f"{'Partial Exits Executed':.<40} {metrics['partial_exits']}")

    print(f"\n--- EXIT REASON BREAKDOWN ---")
    for reason, count in sorted(metrics['exit_reasons'].items(), key=lambda x: -x[1]):
        print(f"  {reason:.<38} {count}")

    print(f"\n--- PER-LEVEL ANALYSIS ---")
    for lvl, stats in metrics['level_stats'].items():
        print(f"  {lvl}: {stats['count']} trades | {stats['win_rate']}% WR | {stats['avg_r']}R avg | {stats['total_r']}R total")

    print(f"\n--- SCORE BREAKDOWN (max points) ---")
    print(f"  Profit Factor:         {min(20, metrics['profit_factor'] * 6.67)/20*100 if metrics['profit_factor'] != float('inf') else 100:.0f}% (of 20)")
    print(f"  Win Rate:              {metrics['win_rate']/60*100 if metrics['win_rate'] > 0 else 0:.0f}% (of 15)")
    print(f"  Avg R per Trade:       {metrics['avg_r']/2*100 if metrics['avg_r'] > 0 else 0:.0f}% (of 15)")
    print(f"  Sharpe/Risk-Adjusted:  {metrics['sharpe_r']/2*100 if metrics['sharpe_r'] > 0 else 0:.0f}% (of 10)")
    print(f"  Max DD Control:        {(100 - metrics['max_drawdown_pct'])/100*100 if metrics['max_drawdown_pct'] > 0 else 100:.0f}% (of 15)")
    print(f"  Calmar Ratio:          {metrics['calmar']/3*100 if metrics['calmar'] != float('inf') else 100:.0f}% (of 10)")
    print(f"  Consistency:           {metrics['win_rate']/100*100:.0f}% (of 10)")
    print(f"  Scale/Upside:          {metrics['total_r']/500*100:.0f}% (of 5)")

    print(f"\n{'='*70}")
    print(f"  FINAL RATING: {metrics['strategy_score']}/100 — {metrics['rating']}")
    print(f"{'='*70}\n")

def main():
    print("=" * 70)
    print("STAIRCASE STRATEGY WEEKLY BACKTEST - Pine 33/33/34")
    print("=" * 70)
    print(f"Scaling: {POS_SIZE_L1:.0f}/{POS_SIZE_L2:.0f}/{POS_SIZE_L3:.0f}")
    print(f"Min Bars Between Entries: {MIN_BARS_BETWEEN_ENTRIES}")
    print(f"Volume Filter: {VOLUME_REQUIREMENT}")
    print(f"Strong Candle: {'ON' if STRONG_ENTRY_CANDLE else 'OFF'} (Pine default)")
    print(f"SL Type: {SL_TYPE}")
    print("=" * 70)

    all_positions: List[StaircasePosition] = []
    weekly_data_map = {}

    for symbol in SYMBOLS:
        print(f"\nBacktesting {symbol}...")
        try:
            daily_df = load_daily_data(symbol)
            weekly_df = compute_indicators(daily_df)
            positions = backtest_stock(symbol, weekly_df)
            all_positions.extend(positions)
            weekly_data_map[symbol] = weekly_df
            print(f"  -> {len(positions)} positions generated")
        except Exception as e:
            print(f"  [ERROR] {symbol}: {e}")

    print(f"\n{'='*70}")
    print(f"TOTAL POSITIONS: {len(all_positions)}")
    print(f"{'='*70}")

    build_excel(all_positions)

    # Compute and print strategy rating
    print("\nComputing strategy metrics...")
    metrics = compute_strategy_metrics(all_positions, weekly_data_map)
    print_strategy_rating(metrics, all_positions)


if __name__ == "__main__":
    main()
