#!/usr/bin/env python3
"""
Staircase Weekly + Minervini Trade Reconciler
===============================================
Analyzes Minervini since-inception trades against Staircase strategy
running on WEEKLY timeframe with the settings from user screenshots.

Settings from screenshots:
- Chart timeframe: Weekly
- Stop Loss: Consolidation Low, 0.2 ATR buffer
- Take Profit: Risk Reward, 2R for L1/L2/L3
- Strong Entry Candle: OFF
- Volume Filter: ON, Above Any (10/20/30)
- Timeframe Filters: ON
  - TF1: Weekly, Short-Term Trend ON, Long-Term OFF, Darvas OFF
  - TF2: Daily, Short-Term Trend ON, Long-Term OFF, Darvas OFF
- Global Darvas 1: Weekly, 40 lookback candles

Outputs:
- Excel workbook with:
  1. Trade_Log     : All Minervini trades + Staircase weekly context at entry
  2. Weekly_Detail : Per-week indicator values for each stock
  3. Signals_In_Trade: Add-on signals detected during each Minervini trade
"""
from __future__ import annotations

import json
import math
import warnings
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional, Dict, Tuple, Any

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")

# =============================================================================
# PATHS
# =============================================================================
ROOT = Path(__file__).resolve().parents[1]
BATCH_JSON = ROOT / "analysis" / "minervini_obsidian" / "data" / "minervini_batch_20260520.json"
DATA_DHAN_DAILY = ROOT / "AP" / "analysis" / "fresh_dhan_since_inception_runs" / "fresh_user_token_20260502_v2" / "raw_dhan_daily"
DATA_CHRONOS = ROOT / "analysis" / "dhan_chronos_backtest" / "cache"
DATA_OHLCV = ROOT / "analysis" / "minervini_obsidian" / "data" / "ohlcv"
OUTPUT_DIR = ROOT / "analysis" / "staircase_dhan"
OUTPUT_EXCEL = OUTPUT_DIR / "staircase_weekly_minervini_reconciliation.xlsx"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# =============================================================================
# CONFIGURATION (from Pine Script screenshots)
# =============================================================================
CHART_TF = "W"           # Weekly chart
MIN_BARS_BETWEEN_ENTRIES = 3
MIN_PROFIT_FOR_NEXT_ENTRY = 0.0   # 0% = breakeven
RISK_REWARD_L1 = 2.0
RISK_REWARD_L2 = 2.0
RISK_REWARD_L3 = 2.0
PARTIAL_EXIT_L1 = 50.0
PARTIAL_EXIT_L2 = 50.0
PARTIAL_EXIT_L3 = 50.0
SL_BUFFER_ATR = 0.2
SL_TYPE = "Consolidation Low"
TP_MODE = "Risk Reward"
STRONG_ENTRY_CANDLE = False
VOLUME_FILTER_ON = True
VOLUME_REQUIREMENT = "Above Any (10/20/30)"

# Timeframe Filter settings
TF1_ON = True
TF1_TF = "W"
TF1_SHORT_TREND = True
TF1_LONG_TREND = False
TF1_DARVAS = False

TF2_ON = True
TF2_TF = "D"
TF2_SHORT_TREND = True
TF2_LONG_TREND = False
TF2_DARVAS = False

# Global Darvas
GLOBAL_DARVAS1_ON = True
GLOBAL_DARVAS1_TF = "W"
GLOBAL_DARVAS1_LEN = 40
GLOBAL_DARVAS2_ON = False

# Slope lookback
SLOPE_LEN = 5

# EMA / SMA lengths
EMA1_LEN = 9
EMA2_LEN = 21
SMA10_LEN = 10
SMA20_LEN = 20
SMA50_LEN = 50
SMA100_LEN = 100
SMA200_LEN = 200

ATR_LEN = 14

# =============================================================================
# DATA LOADING
# =============================================================================

def load_daily_data(symbol: str) -> pd.DataFrame:
    """Load daily OHLCV for a symbol from local caches or yfinance."""
    symbol_upper = symbol.upper()
    df = None

    # 1. Dhan since-inception daily
    dhan_path = DATA_DHAN_DAILY / f"{symbol_upper}_daily_1990-01-01_2026-05-02.csv"
    if dhan_path.exists():
        df = pd.read_csv(dhan_path)
        df["date"] = pd.to_datetime(df["date"])
        df = df.sort_values("date").reset_index(drop=True)
        df.columns = [c.lower() for c in df.columns]

    # 2. Obsidian OHLCV
    if df is None:
        obs_path = DATA_OHLCV / f"{symbol_upper}.csv"
        if obs_path.exists():
            df = pd.read_csv(obs_path)
            df["Date"] = pd.to_datetime(df["Date"])
            df = df.sort_values("Date").reset_index(drop=True)
            df = df.rename(columns={
                "Date": "date", "Open": "open", "High": "high",
                "Low": "low", "Close": "close", "Volume": "volume"
            })

    # 3. Dhan chronos daily
    if df is None:
        for suffix in ["_D_2021-05-13_2026-05-13.csv", "_D_2022-01-01_2026-05-13.csv"]:
            chrono_path = DATA_CHRONOS / f"{symbol_upper}{suffix}"
            if chrono_path.exists():
                dfr = pd.read_csv(chrono_path)
                dfr["datetime"] = pd.to_datetime(dfr["datetime"])
                dfr = dfr.sort_values("datetime").reset_index(drop=True)
                dfr["date"] = dfr["datetime"].dt.date
                daily = dfr.groupby("date").agg({
                    "open": "first", "high": "max", "low": "min",
                    "close": "last", "volume": "sum"
                }).reset_index()
                daily["date"] = pd.to_datetime(daily["date"])
                df = daily
                break

    # 4. yfinance fallback
    if df is None:
        for suffix in [".NS", ".BO"]:
            try:
                ticker = yf.Ticker(f"{symbol_upper}{suffix}")
                hist = ticker.history(period="max", auto_adjust=False)
                if len(hist) == 0:
                    continue
                hist = hist.reset_index()
                hist.columns = [c.lower().replace(" ", "_") for c in hist.columns]
                if isinstance(hist.columns, pd.MultiIndex):
                    hist.columns = hist.columns.get_level_values(0)
                if "date" not in hist.columns:
                    hist = hist.rename(columns={hist.columns[0]: "date"})
                hist["date"] = pd.to_datetime(hist["date"]).dt.tz_localize(None)
                df = hist[["date", "open", "high", "low", "close", "volume"]].copy()
                if len(df) > 50:
                    break
            except Exception as e:
                print(f"  [WARN] yfinance {suffix} failed for {symbol}: {e}")

    if df is None or len(df) < 50:
        raise ValueError(f"No usable daily data for {symbol}")

    for col in ["open", "high", "low", "close", "volume"]:
        if col not in df.columns:
            raise ValueError(f"Missing column '{col}' in daily data for {symbol}")

    df = df.sort_values("date").reset_index(drop=True)
    return df


# =============================================================================
# INDICATOR COMPUTATIONS
# =============================================================================

def compute_emas(df: pd.DataFrame, prefix: str = "") -> pd.DataFrame:
    """Compute EMAs with standard lengths."""
    df = df.copy()
    df[f"{prefix}ema9"] = df["close"].ewm(span=EMA1_LEN, adjust=False).mean()
    df[f"{prefix}ema21"] = df["close"].ewm(span=EMA2_LEN, adjust=False).mean()
    df[f"{prefix}ema10"] = df["close"].ewm(span=SMA10_LEN, adjust=False).mean()
    df[f"{prefix}ema20"] = df["close"].ewm(span=SMA20_LEN, adjust=False).mean()
    df[f"{prefix}ema50"] = df["close"].ewm(span=SMA50_LEN, adjust=False).mean()
    df[f"{prefix}ema100"] = df["close"].ewm(span=SMA100_LEN, adjust=False).mean()
    df[f"{prefix}ema200"] = df["close"].ewm(span=SMA200_LEN, adjust=False).mean()
    return df


def compute_smas(df: pd.DataFrame, prefix: str = "") -> pd.DataFrame:
    """Compute SMAs with standard lengths."""
    df = df.copy()
    df[f"{prefix}sma10"] = df["close"].rolling(SMA10_LEN).mean()
    df[f"{prefix}sma20"] = df["close"].rolling(SMA20_LEN).mean()
    df[f"{prefix}sma50"] = df["close"].rolling(SMA50_LEN).mean()
    df[f"{prefix}sma100"] = df["close"].rolling(SMA100_LEN).mean()
    df[f"{prefix}sma200"] = df["close"].rolling(SMA200_LEN).mean()
    return df


def compute_atr(df: pd.DataFrame, length: int = ATR_LEN, prefix: str = "") -> pd.DataFrame:
    """Compute ATR."""
    df = df.copy()
    prev_close = df["close"].shift(1)
    tr1 = df["high"] - df["low"]
    tr2 = (df["high"] - prev_close).abs()
    tr3 = (df["low"] - prev_close).abs()
    df[f"{prefix}tr"] = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
    df[f"{prefix}atr"] = df[f"{prefix}tr"].rolling(length).mean()
    return df


def compute_volume_ok(df: pd.DataFrame, prefix: str = "") -> pd.DataFrame:
    """Compute volume filter: Above Any (10/20/30)."""
    df = df.copy()
    df[f"{prefix}vol10"] = df["volume"].rolling(10).mean()
    df[f"{prefix}vol20"] = df["volume"].rolling(20).mean()
    df[f"{prefix}vol30"] = df["volume"].rolling(30).mean()
    above10 = df["volume"] > df[f"{prefix}vol10"]
    above20 = df["volume"] > df[f"{prefix}vol20"]
    above30 = df["volume"] > df[f"{prefix}vol30"]
    df[f"{prefix}volume_ok"] = above10 | above20 | above30
    return df


def compute_darvas(df: pd.DataFrame, lookback: int, prefix: str = "") -> pd.DataFrame:
    """Compute Darvas box using body tops/bottoms."""
    df = df.copy()
    body_top = df[["open", "close"]].max(axis=1)
    body_bottom = df[["open", "close"]].min(axis=1)
    df[f"{prefix}darvas_top"] = body_top.rolling(lookback).max().shift(1)
    df[f"{prefix}darvas_bottom"] = body_bottom.rolling(lookback).min().shift(1)
    df[f"{prefix}darvas_ok"] = df["close"] > df[f"{prefix}darvas_top"]
    return df


def compute_weekly_from_daily(daily_df: pd.DataFrame) -> pd.DataFrame:
    """Resample daily OHLCV to weekly (Friday-based)."""
    daily_df = daily_df.copy()
    daily_df = daily_df.set_index("date")
    weekly = daily_df.resample("W-FRI").agg({
        "open": "first",
        "high": "max",
        "low": "min",
        "close": "last",
        "volume": "sum"
    }).dropna().reset_index()
    weekly = weekly.sort_values("date").reset_index(drop=True)
    return weekly


def compute_staircase_weekly(daily_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Compute all Staircase indicators on weekly (chart) and daily (TF2) timeframes.
    Returns (weekly_df with all indicators, daily_df with all indicators).
    """
    # --- Daily indicators (for TF2) ---
    daily = daily_df.copy()
    daily = compute_emas(daily, prefix="d_")
    daily = compute_smas(daily, prefix="d_")
    daily = compute_volume_ok(daily, prefix="d_")

    # Daily short-term trend (EMA-based as in Pine calcTFData)
    daily["d_short_term_stacked"] = (
        (daily["d_ema10"] > daily["d_ema20"]) &
        (daily["d_ema20"] > daily["d_ema50"])
    )
    daily["d_price_above_short_emas"] = (
        (daily["close"] > daily["d_ema10"]) &
        (daily["close"] > daily["d_ema20"]) &
        (daily["close"] > daily["d_ema50"])
    )
    daily["d_short_term_trend_ok"] = daily["d_short_term_stacked"] & daily["d_price_above_short_emas"]

    # Daily long-term trend (not used since TF2 Long-Term is OFF, but compute anyway)
    daily["d_long_term_stacked"] = (
        (daily["d_ema50"] > daily["d_ema100"]) &
        (daily["d_ema100"] > daily["d_ema200"])
    )
    daily["d_ema50_rising"] = daily["d_ema50"] > daily["d_ema50"].shift(SLOPE_LEN)
    daily["d_ema100_rising"] = daily["d_ema100"] > daily["d_ema100"].shift(SLOPE_LEN)
    daily["d_price_above_long_emas"] = (
        (daily["close"] > daily["d_ema50"]) &
        (daily["close"] > daily["d_ema100"]) &
        (daily["close"] > daily["d_ema200"])
    )
    daily["d_major_trend_ok"] = (
        daily["d_long_term_stacked"] &
        daily["d_ema50_rising"] &
        daily["d_ema100_rising"] &
        daily["d_price_above_long_emas"]
    )

    # Daily EMA uptrend
    daily["d_ema_uptrend"] = daily["d_ema9"] > daily["d_ema21"]
    daily["d_price_above_ema"] = daily["close"] > daily["d_ema21"]

    # Daily TF2 signal = Darvas OK (OFF) & Short OK & Long OK (OFF) & EMA OK & Volume OK
    daily["d_tf2_short_ok"] = ~TF2_SHORT_TREND | daily["d_short_term_trend_ok"]
    daily["d_tf2_long_ok"] = ~TF2_LONG_TREND | daily["d_major_trend_ok"]
    daily["d_tf2_ema_ok"] = daily["d_ema_uptrend"] & daily["d_price_above_ema"]
    daily["d_tf2_ok"] = daily["d_tf2_short_ok"] & daily["d_tf2_long_ok"] & daily["d_tf2_ema_ok"] & daily["d_volume_ok"]

    # --- Weekly indicators (chart timeframe) ---
    weekly = compute_weekly_from_daily(daily_df)
    weekly = compute_emas(weekly, prefix="w_")
    weekly = compute_smas(weekly, prefix="w_")
    weekly = compute_atr(weekly, prefix="w_")
    weekly = compute_volume_ok(weekly, prefix="w_")
    weekly = compute_darvas(weekly, GLOBAL_DARVAS1_LEN, prefix="w_gd1_")

    # Weekly short-term trend (EMA-based)
    weekly["w_short_term_stacked"] = (
        (weekly["w_ema10"] > weekly["w_ema20"]) &
        (weekly["w_ema20"] > weekly["w_ema50"])
    )
    weekly["w_price_above_short_emas"] = (
        (weekly["close"] > weekly["w_ema10"]) &
        (weekly["close"] > weekly["w_ema20"]) &
        (weekly["close"] > weekly["w_ema50"])
    )
    weekly["w_short_term_trend_ok"] = weekly["w_short_term_stacked"] & weekly["w_price_above_short_emas"]

    # Weekly long-term trend
    weekly["w_long_term_stacked"] = (
        (weekly["w_ema50"] > weekly["w_ema100"]) &
        (weekly["w_ema100"] > weekly["w_ema200"])
    )
    weekly["w_ema50_rising"] = weekly["w_ema50"] > weekly["w_ema50"].shift(SLOPE_LEN)
    weekly["w_ema100_rising"] = weekly["w_ema100"] > weekly["w_ema100"].shift(SLOPE_LEN)
    weekly["w_ema200_rising"] = weekly["w_ema200"] > weekly["w_ema200"].shift(SLOPE_LEN)
    weekly["w_core_rising"] = weekly["w_ema50_rising"] & weekly["w_ema100_rising"] & weekly["w_ema200_rising"]
    weekly["w_price_above_long_emas"] = (
        (weekly["close"] > weekly["w_ema50"]) &
        (weekly["close"] > weekly["w_ema100"]) &
        (weekly["close"] > weekly["w_ema200"])
    )
    weekly["w_major_trend_ok"] = (
        weekly["w_long_term_stacked"] &
        weekly["w_ema50_rising"] &
        weekly["w_ema100_rising"] &
        weekly["w_price_above_long_emas"]
    )

    # Weekly full_bull = short stacked + long stacked + price above all
    weekly["w_full_bull"] = (
        weekly["w_short_term_stacked"] &
        weekly["w_long_term_stacked"] &
        weekly["w_price_above_short_emas"] &
        weekly["w_price_above_long_emas"]
    )

    # Weekly EMA uptrend (chart base)
    weekly["w_ema_uptrend"] = weekly["w_ema9"] > weekly["w_ema21"]
    weekly["w_price_above_ema"] = weekly["close"] > weekly["w_ema21"]

    # Weekly base entry core
    weekly["w_base_entry_core"] = (
        weekly["w_ema_uptrend"] &
        weekly["w_price_above_ema"] &
        weekly["w_volume_ok"]
    )
    if STRONG_ENTRY_CANDLE:
        candle_range = weekly["high"] - weekly["low"]
        body_range = (weekly["close"] - weekly["open"]).abs()
        body_pct = np.where(candle_range > 0, body_range / candle_range * 100, 0)
        close_from_high = np.where(candle_range > 0, (weekly["high"] - weekly["close"]) / candle_range * 100, 100)
        strong = (weekly["close"] > weekly["open"]) & (body_pct >= 40) & (close_from_high <= 25)
        weekly["w_base_entry_core"] = weekly["w_base_entry_core"] & strong

    # Weekly TF1 signal (same timeframe as chart, but calcTFData logic)
    weekly["w_tf1_short_ok"] = ~TF1_SHORT_TREND | weekly["w_short_term_trend_ok"]
    weekly["w_tf1_long_ok"] = ~TF1_LONG_TREND | weekly["w_major_trend_ok"]
    weekly["w_tf1_ema_ok"] = weekly["w_ema_uptrend"] & weekly["w_price_above_ema"]
    weekly["w_tf1_ok"] = weekly["w_tf1_short_ok"] & weekly["w_tf1_long_ok"] & weekly["w_tf1_ema_ok"] & weekly["w_volume_ok"]

    # Global Darvas 1
    weekly["w_global_darvas1_ok"] = ~GLOBAL_DARVAS1_ON | weekly["w_gd1_darvas_ok"]

    # Map daily TF2 values to weekly bars (use Friday's daily value)
    # For each weekly date, find the daily row with the same or last preceding date
    daily_last = daily.copy()
    daily_last = daily_last.set_index("date")
    weekly = weekly.set_index("date")

    # Reindex daily to weekly dates (forward fill)
    for col in ["d_tf2_ok", "d_ema_uptrend", "d_short_term_trend_ok", "d_price_above_ema", "d_volume_ok",
                "d_ema9", "d_ema21", "d_ema50"]:
        s = daily_last[col]
        weekly[f"mapped_{col}"] = s.reindex(weekly.index, method="ffill")

    weekly = weekly.reset_index()

    # Full entry signal (weekly chart)
    weekly["w_entry_signal_l1"] = (
        weekly["w_base_entry_core"] &
        weekly["w_tf1_ok"] &
        weekly["mapped_d_tf2_ok"] &
        weekly["w_global_darvas1_ok"]
    )

    return weekly, daily


# =============================================================================
# TRADE RECONCILIATION
# =============================================================================

@dataclass
class StaircaseContext:
    """Staircase weekly context for a single Minervini trade."""
    # Weekly at entry
    w_date: Optional[str] = None
    w_close: float = 0.0
    w_ema9: float = 0.0
    w_ema21: float = 0.0
    w_ema50: float = 0.0
    w_ema100: float = 0.0
    w_ema200: float = 0.0
    w_ema_uptrend: bool = False
    w_short_term_stacked: bool = False
    w_price_above_short_emas: bool = False
    w_long_term_stacked: bool = False
    w_core_rising: bool = False
    w_price_above_long_emas: bool = False
    w_full_bull: bool = False
    w_major_trend_ok: bool = False
    w_tf1_ok: bool = False
    w_global_darvas_top: float = 0.0
    w_global_darvas_ok: bool = False
    w_base_entry_core: bool = False
    w_entry_signal: bool = False

    # Daily (Friday) at entry
    d_tf2_ok: bool = False
    d_ema_uptrend: bool = False
    d_short_term_trend_ok: bool = False

    # During trade
    weeks_in_trade: int = 0
    weeks_structure_aligned: int = 0
    weeks_structure_broken: int = 0
    potential_l2_signals: int = 0
    potential_l3_signals: int = 0
    structure_broke_then_recovered: bool = False
    first_break_week: Optional[str] = None
    recovery_week: Optional[str] = None

    # Exit context
    exit_w_full_bull: bool = False
    exit_w_major_trend_ok: bool = False
    exit_w_ema_uptrend: bool = False


def reconcile_trade(
    symbol: str,
    trade: Dict[str, Any],
    weekly_df: pd.DataFrame,
    daily_df: pd.DataFrame
) -> StaircaseContext:
    """Reconcile a single Minervini trade against Staircase weekly strategy."""
    ctx = StaircaseContext()

    entry_date = pd.to_datetime(trade["entry_date"])
    exit_date = pd.to_datetime(trade["exit_date"]) if trade.get("exit_date") else None

    # Find entry week (last Friday <= entry_date, or next Friday if entry on weekend)
    wk = weekly_df.copy()
    entry_mask = wk["date"] >= entry_date
    if entry_mask.any():
        entry_row = wk[entry_mask].iloc[0]
    else:
        entry_row = wk.iloc[-1]

    ctx.w_date = entry_row["date"].strftime("%Y-%m-%d")
    ctx.w_close = round(entry_row["close"], 2)
    ctx.w_ema9 = round(entry_row["w_ema9"], 2)
    ctx.w_ema21 = round(entry_row["w_ema21"], 2)
    ctx.w_ema50 = round(entry_row["w_ema50"], 2)
    ctx.w_ema100 = round(entry_row["w_ema100"], 2)
    ctx.w_ema200 = round(entry_row["w_ema200"], 2)
    ctx.w_ema_uptrend = bool(entry_row["w_ema_uptrend"])
    ctx.w_short_term_stacked = bool(entry_row["w_short_term_stacked"])
    ctx.w_price_above_short_emas = bool(entry_row["w_price_above_short_emas"])
    ctx.w_long_term_stacked = bool(entry_row["w_long_term_stacked"])
    ctx.w_core_rising = bool(entry_row["w_core_rising"])
    ctx.w_price_above_long_emas = bool(entry_row["w_price_above_long_emas"])
    ctx.w_full_bull = bool(entry_row["w_full_bull"])
    ctx.w_major_trend_ok = bool(entry_row["w_major_trend_ok"])
    ctx.w_tf1_ok = bool(entry_row["w_tf1_ok"])
    ctx.w_global_darvas_top = round(entry_row["w_gd1_darvas_top"], 2) if pd.notna(entry_row["w_gd1_darvas_top"]) else 0.0
    ctx.w_global_darvas_ok = bool(entry_row["w_global_darvas1_ok"])
    ctx.w_base_entry_core = bool(entry_row["w_base_entry_core"])
    ctx.w_entry_signal = bool(entry_row["w_entry_signal_l1"])

    ctx.d_tf2_ok = bool(entry_row["mapped_d_tf2_ok"]) if pd.notna(entry_row["mapped_d_tf2_ok"]) else False
    ctx.d_ema_uptrend = bool(entry_row["mapped_d_ema_uptrend"]) if pd.notna(entry_row["mapped_d_ema_uptrend"]) else False
    ctx.d_short_term_trend_ok = bool(entry_row["mapped_d_short_term_trend_ok"]) if pd.notna(entry_row["mapped_d_short_term_trend_ok"]) else False

    # --- Analyze holding period ---
    if exit_date is not None:
        period_mask = (wk["date"] >= entry_row["date"]) & (wk["date"] <= exit_date)
    else:
        period_mask = wk["date"] >= entry_row["date"]

    period = wk[period_mask].copy()
    ctx.weeks_in_trade = len(period)

    if len(period) > 0:
        # Structure alignment during trade
        aligned = period["w_full_bull"] | period["w_major_trend_ok"]
        ctx.weeks_structure_aligned = int(aligned.sum())
        ctx.weeks_structure_broken = len(period) - ctx.weeks_structure_aligned

        # Check for break then recovery
        broke = (~aligned).any()
        if broke:
            first_break = period[~aligned].iloc[0]["date"]
            ctx.first_break_week = first_break.strftime("%Y-%m-%d")
            after_break = period[period["date"] > first_break]
            if len(after_break) > 0 and (after_break["w_full_bull"] | after_break["w_major_trend_ok"]).any():
                recovery = after_break[after_break["w_full_bull"] | after_break["w_major_trend_ok"]].iloc[0]["date"]
                ctx.recovery_week = recovery.strftime("%Y-%m-%d")
                ctx.structure_broke_then_recovered = True

        # Simulate staircase add-ons during trade
        # L2: need L1 in profit (close > entry1_price), base + filters OK, >=3 weeks since L1
        # L3: need L2 in profit, base + filters OK, >=3 weeks since L2
        # Since Minervini is one position, we simulate staircase pyramiding
        entry_price = trade["entry_price"]

        # Skip first week (L1 entry week)
        addon_candidates = period.iloc[1:].copy()
        if len(addon_candidates) >= MIN_BARS_BETWEEN_ENTRIES:
            l1_active = True
            l1_entry_week = period.iloc[0]["date"]
            l2_active = False
            l2_entry_week = None

            for _, row in addon_candidates.iterrows():
                weeks_since_l1 = (row["date"] - l1_entry_week).days / 7
                in_profit = row["close"] > entry_price
                signal = bool(row["w_entry_signal_l1"])

                if l1_active and not l2_active and weeks_since_l1 >= MIN_BARS_BETWEEN_ENTRIES and in_profit and signal:
                    ctx.potential_l2_signals += 1
                    l2_active = True
                    l2_entry_week = row["date"]

                if l2_active:
                    weeks_since_l2 = (row["date"] - l2_entry_week).days / 7
                    if weeks_since_l2 >= MIN_BARS_BETWEEN_ENTRIES and in_profit and signal:
                        ctx.potential_l3_signals += 1

        # Exit week context
        if exit_date is not None and len(period) > 0:
            exit_row = period.iloc[-1]
            ctx.exit_w_full_bull = bool(exit_row["w_full_bull"])
            ctx.exit_w_major_trend_ok = bool(exit_row["w_major_trend_ok"])
            ctx.exit_w_ema_uptrend = bool(exit_row["w_ema_uptrend"])

    return ctx


# =============================================================================
# EXCEL BUILDER
# =============================================================================

def build_excel(trade_rows: List[Dict], weekly_detail_rows: List[Dict], signal_rows: List[Dict]):
    """Build the Excel workbook with formatted sheets."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # --- Sheet 1: Trade Log ---
    ws1 = wb.create_sheet("Trade_Log")
    df_trades = pd.DataFrame(trade_rows)
    for r_idx, row in enumerate(dataframe_to_rows(df_trades, index=False, header=True), 1):
        ws1.append(row)
    # Header formatting
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    # Auto-width
    for col in ws1.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws1.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # Freeze header
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Weekly_Detail ---
    ws2 = wb.create_sheet("Weekly_Detail")
    df_weekly = pd.DataFrame(weekly_detail_rows)
    for r_idx, row in enumerate(dataframe_to_rows(df_weekly, index=False, header=True), 1):
        ws2.append(row)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws2.column_dimensions[col_letter].width = min(max_len + 2, 40)
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Signals_In_Trade ---
    ws3 = wb.create_sheet("Signals_In_Trade")
    df_signals = pd.DataFrame(signal_rows)
    if df_signals.empty:
        df_signals = pd.DataFrame(columns=[
            "symbol", "trade_entry", "trade_exit", "signal_week", "signal_type",
            "week_close", "entry_price", "pnl_pct", "reason"
        ])
    for r_idx, row in enumerate(dataframe_to_rows(df_signals, index=False, header=True), 1):
        ws3.append(row)
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col in ws3.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws3.column_dimensions[col_letter].width = min(max_len + 2, 40)
    ws3.freeze_panes = "A2"

    wb.save(OUTPUT_EXCEL)
    print(f"\n[OK] Excel saved: {OUTPUT_EXCEL}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 70)
    print("STAIRCASE WEEKLY + MINERVINI TRADE RECONCILER")
    print("=" * 70)

    # Load batch
    with open(BATCH_JSON, "r") as f:
        batch = json.load(f)

    print(f"Loaded batch with {len(batch)} stocks\n")

    trade_rows = []
    weekly_detail_rows = []
    signal_rows = []

    for stock_data in batch:
        symbol = stock_data["symbol"]
        trades = stock_data.get("trades", [])
        print(f"Processing {symbol} ({len(trades)} trades)...")

        try:
            daily_df = load_daily_data(symbol)
            weekly_df, daily_df_ind = compute_staircase_weekly(daily_df)
        except Exception as e:
            print(f"  [ERROR] {symbol}: {e}")
            continue

        # Build weekly detail for this stock
        for _, row in weekly_df.iterrows():
            weekly_detail_rows.append({
                "symbol": symbol,
                "week_date": row["date"].strftime("%Y-%m-%d"),
                "open": round(row["open"], 2),
                "high": round(row["high"], 2),
                "low": round(row["low"], 2),
                "close": round(row["close"], 2),
                "volume": int(row["volume"]),
                "w_ema9": round(row["w_ema9"], 2),
                "w_ema21": round(row["w_ema21"], 2),
                "w_ema50": round(row["w_ema50"], 2),
                "w_ema100": round(row["w_ema100"], 2),
                "w_ema200": round(row["w_ema200"], 2),
                "w_ema_uptrend": bool(row["w_ema_uptrend"]),
                "w_short_term_stacked": bool(row["w_short_term_stacked"]),
                "w_long_term_stacked": bool(row["w_long_term_stacked"]),
                "w_core_rising": bool(row["w_core_rising"]),
                "w_full_bull": bool(row["w_full_bull"]),
                "w_major_trend_ok": bool(row["w_major_trend_ok"]),
                "w_base_entry_core": bool(row["w_base_entry_core"]),
                "w_tf1_ok": bool(row["w_tf1_ok"]),
                "w_global_darvas_top": round(row["w_gd1_darvas_top"], 2) if pd.notna(row["w_gd1_darvas_top"]) else None,
                "w_global_darvas_ok": bool(row["w_global_darvas1_ok"]),
                "w_entry_signal": bool(row["w_entry_signal_l1"]),
                "d_tf2_ok": bool(row["mapped_d_tf2_ok"]) if pd.notna(row["mapped_d_tf2_ok"]) else None,
            })

        for trade in trades:
            ctx = reconcile_trade(symbol, trade, weekly_df, daily_df_ind)

            # Determine if Minervini trade aligned with staircase
            staircase_aligned = ctx.w_entry_signal
            trend_aligned = ctx.w_full_bull or ctx.w_major_trend_ok

            trade_rows.append({
                "symbol": symbol,
                "minervini_entry": trade["entry_date"],
                "minervini_entry_price": trade["entry_price"],
                "minervini_exit": trade.get("exit_date", "OPEN"),
                "minervini_exit_price": trade.get("exit_price", ""),
                "exit_reason": trade.get("exit_reason", ""),
                "days_held": trade.get("days_held", 0),
                "r_multiple_full": round(trade.get("r_multiple_full", 0), 2),
                "r_multiple_weighted": round(trade.get("r_multiple_weighted", 0), 2),
                "week_of_entry": ctx.w_date,
                "w_close": ctx.w_close,
                "w_ema9": ctx.w_ema9,
                "w_ema21": ctx.w_ema21,
                "w_ema50": ctx.w_ema50,
                "w_ema100": ctx.w_ema100,
                "w_ema200": ctx.w_ema200,
                "w_ema_uptrend": ctx.w_ema_uptrend,
                "w_short_term_stacked": ctx.w_short_term_stacked,
                "w_price_above_short": ctx.w_price_above_short_emas,
                "w_long_term_stacked": ctx.w_long_term_stacked,
                "w_core_rising": ctx.w_core_rising,
                "w_price_above_long": ctx.w_price_above_long_emas,
                "w_full_bull": ctx.w_full_bull,
                "w_major_trend_ok": ctx.w_major_trend_ok,
                "w_tf1_pass": ctx.w_tf1_ok,
                "w_darvas_top_40w": ctx.w_global_darvas_top,
                "w_global_darvas_pass": ctx.w_global_darvas_ok,
                "w_base_entry_core": ctx.w_base_entry_core,
                "d_tf2_pass": ctx.d_tf2_ok,
                "d_ema_uptrend": ctx.d_ema_uptrend,
                "d_short_term_ok": ctx.d_short_term_trend_ok,
                "staircase_would_enter": ctx.w_entry_signal,
                "trend_aligned_at_entry": trend_aligned,
                "weeks_in_trade": ctx.weeks_in_trade,
                "weeks_structure_aligned": ctx.weeks_structure_aligned,
                "weeks_structure_broken": ctx.weeks_structure_broken,
                "potential_l2_signals": ctx.potential_l2_signals,
                "potential_l3_signals": ctx.potential_l3_signals,
                "structure_broke_then_recovered": ctx.structure_broke_then_recovered,
                "first_break_week": ctx.first_break_week or "",
                "recovery_week": ctx.recovery_week or "",
                "exit_w_full_bull": ctx.exit_w_full_bull,
                "exit_w_major_trend_ok": ctx.exit_w_major_trend_ok,
                "exit_w_ema_uptrend": ctx.exit_w_ema_uptrend,
            })

            # Signal-in-trade detail rows
            entry_date = pd.to_datetime(trade["entry_date"])
            exit_date = pd.to_datetime(trade["exit_date"]) if trade.get("exit_date") else None
            entry_price = trade["entry_price"]

            wk = weekly_df.copy()
            if exit_date is not None:
                period_mask = (wk["date"] >= entry_date) & (wk["date"] <= exit_date)
            else:
                period_mask = wk["date"] >= entry_date
            period = wk[period_mask].copy()

            # Skip first week (L1)
            addon_period = period.iloc[1:].copy() if len(period) > 1 else pd.DataFrame()

            l1_active = True
            l1_entry_week = period.iloc[0]["date"] if len(period) > 0 else None
            l2_active = False
            l2_entry_week = None

            for _, row in addon_period.iterrows():
                if l1_entry_week is None:
                    break
                weeks_since_l1 = (row["date"] - l1_entry_week).days / 7
                in_profit = row["close"] > entry_price
                signal = bool(row["w_entry_signal_l1"])

                if l1_active and not l2_active and weeks_since_l1 >= MIN_BARS_BETWEEN_ENTRIES and in_profit and signal:
                    signal_rows.append({
                        "symbol": symbol,
                        "trade_entry": trade["entry_date"],
                        "trade_exit": trade.get("exit_date", "OPEN"),
                        "signal_week": row["date"].strftime("%Y-%m-%d"),
                        "signal_type": "L2_ADD_ON",
                        "week_close": round(row["close"], 2),
                        "entry_price": entry_price,
                        "pnl_pct": round((row["close"] - entry_price) / entry_price * 100, 2),
                        "reason": f"Weekly entry signal at +{round((row['close'] - entry_price) / entry_price * 100, 1)}% profit, {int(weeks_since_l1)}w since entry"
                    })
                    l2_active = True
                    l2_entry_week = row["date"]

                if l2_active and l2_entry_week is not None:
                    weeks_since_l2 = (row["date"] - l2_entry_week).days / 7
                    if weeks_since_l2 >= MIN_BARS_BETWEEN_ENTRIES and in_profit and signal:
                        signal_rows.append({
                            "symbol": symbol,
                            "trade_entry": trade["entry_date"],
                            "trade_exit": trade.get("exit_date", "OPEN"),
                            "signal_week": row["date"].strftime("%Y-%m-%d"),
                            "signal_type": "L3_ADD_ON",
                            "week_close": round(row["close"], 2),
                            "entry_price": entry_price,
                            "pnl_pct": round((row["close"] - entry_price) / entry_price * 100, 2),
                            "reason": f"Weekly entry signal at +{round((row['close'] - entry_price) / entry_price * 100, 1)}% profit, {int(weeks_since_l2)}w since L2"
                        })

    print(f"\n{'='*70}")
    print(f"Summary:")
    print(f"  Total Minervini trades: {len(trade_rows)}")
    print(f"  Total weekly bars tracked: {len(weekly_detail_rows)}")
    print(f"  Total add-on signals detected: {len(signal_rows)}")
    print(f"{'='*70}")

    build_excel(trade_rows, weekly_detail_rows, signal_rows)


if __name__ == "__main__":
    main()
