#!/usr/bin/env python3
"""
Staircase Strategy Weekly Backtest Engine
==========================================
Full simulation of the Staircase SMA strategy on WEEKLY timeframe
with the exact settings from user screenshots.

Settings:
- Chart TF: Weekly
- SL: Consolidation Low, 0.2 ATR buffer
- TP: Risk Reward, 2R all levels
- Partial: 50% at 2R TP
- Sizing: 33/33/34
- Strong Candle: OFF
- Volume: ON, Above Any (10/20/30)
- TF1: Weekly, Short-Term ON, Long-Term OFF, Darvas OFF
- TF2: Daily, Short-Term ON, Long-Term OFF, Darvas OFF
- Global Darvas 1: Weekly, 40 lookback

Outputs:
- Excel with Trade_Summary, Entry_Log, Partial_Log, Exit_Events, Stock_Summary
"""
from __future__ import annotations

import json
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Tuple

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")

# =============================================================================
# PATHS & CONFIG
# =============================================================================
ROOT = Path(__file__).resolve().parents[1]
DATA_DHAN_DAILY = ROOT / "AP" / "analysis" / "fresh_dhan_since_inception_runs" / "fresh_user_token_20260502_v2" / "raw_dhan_daily"
DATA_CHRONOS = ROOT / "analysis" / "dhan_chronos_backtest" / "cache"
DATA_OHLCV = ROOT / "analysis" / "minervini_obsidian" / "data" / "ohlcv"
OUTPUT_DIR = ROOT / "analysis" / "staircase_dhan"
OUTPUT_EXCEL = OUTPUT_DIR / "staircase_weekly_backtest_trades.xlsx"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Strategy config
MIN_BARS_BETWEEN_ENTRIES = 3
MIN_PROFIT_FOR_NEXT_ENTRY = 0.0
RISK_REWARD = 2.0
PARTIAL_PCT = 50.0
POS_SIZE_L1 = 33.0
POS_SIZE_L2 = 33.0
POS_SIZE_L3 = 34.0
SL_BUFFER_ATR = 0.2
STRONG_ENTRY_CANDLE = False
VOLUME_FILTER_ON = True

TF1_ON = True
TF1_SHORT = True
TF1_LONG = False
TF2_ON = True
TF2_SHORT = True
TF2_LONG = False
GLOBAL_DARVAS1_ON = True
GLOBAL_DARVAS1_LEN = 40
SLOPE_LEN = 5

EMA1_LEN, EMA2_LEN = 9, 21
SMA10_LEN, SMA20_LEN, SMA50_LEN = 10, 20, 50
SMA100_LEN, SMA200_LEN = 100, 200
ATR_LEN = 14

SYMBOLS = ["BSE", "IRFC", "RVNL", "MCX", "COCHINSHIP", "GRSE", "HAL", "BEL", "APOLLOMICRO"]

# =============================================================================
# DATA LOADING (same as reconciler)
# =============================================================================

def load_daily_data(symbol: str) -> pd.DataFrame:
    symbol_upper = symbol.upper()
    df = None

    dhan_path = DATA_DHAN_DAILY / f"{symbol_upper}_daily_1990-01-01_2026-05-02.csv"
    if dhan_path.exists():
        df = pd.read_csv(dhan_path)
        df["date"] = pd.to_datetime(df["date"])
        df = df.sort_values("date").reset_index(drop=True)
        df.columns = [c.lower() for c in df.columns]

    if df is None:
        obs_path = DATA_OHLCV / f"{symbol_upper}.csv"
        if obs_path.exists():
            df = pd.read_csv(obs_path)
            df["Date"] = pd.to_datetime(df["Date"])
            df = df.sort_values("Date").reset_index(drop=True)
            df = df.rename(columns={
                "Date": "date", "Open": "open", "High": "high",
                "Low": "low", "Close": "close", "Volume": "volume"
            })

    if df is None:
        for suffix in ["_D_2021-05-13_2026-05-13.csv", "_D_2022-01-01_2026-05-13.csv"]:
            chrono_path = DATA_CHRONOS / f"{symbol_upper}{suffix}"
            if chrono_path.exists():
                dfr = pd.read_csv(chrono_path)
                dfr["datetime"] = pd.to_datetime(dfr["datetime"])
                dfr = dfr.sort_values("datetime").reset_index(drop=True)
                dfr["date"] = dfr["datetime"].dt.date
                daily = dfr.groupby("date").agg({
                    "open": "first", "high": "max", "low": "min",
                    "close": "last", "volume": "sum"
                }).reset_index()
                daily["date"] = pd.to_datetime(daily["date"])
                df = daily
                break

    if df is None:
        for suffix in [".NS", ".BO"]:
            try:
                ticker = yf.Ticker(f"{symbol_upper}{suffix}")
                hist = ticker.history(period="max", auto_adjust=False)
                if len(hist) == 0:
                    continue
                hist = hist.reset_index()
                hist.columns = [c.lower().replace(" ", "_") for c in hist.columns]
                if isinstance(hist.columns, pd.MultiIndex):
                    hist.columns = hist.columns.get_level_values(0)
                if "date" not in hist.columns:
                    hist = hist.rename(columns={hist.columns[0]: "date"})
                hist["date"] = pd.to_datetime(hist["date"]).dt.tz_localize(None)
                df = hist[["date", "open", "high", "low", "close", "volume"]].copy()
                if len(df) > 50:
                    break
            except Exception:
                pass

    if df is None or len(df) < 50:
        raise ValueError(f"No usable daily data for {symbol}")

    df = df.sort_values("date").reset_index(drop=True)
    return df


def compute_indicators(daily_df: pd.DataFrame) -> pd.DataFrame:
    """Compute weekly + daily indicators and merge into weekly dataframe."""
    # Daily EMAs / SMAs / volume
    daily = daily_df.copy()
    for span, name in [(EMA1_LEN, "ema9"), (EMA2_LEN, "ema21"), (SMA10_LEN, "ema10"),
                       (SMA20_LEN, "ema20"), (SMA50_LEN, "ema50"), (SMA100_LEN, "ema100"), (SMA200_LEN, "ema200")]:
        daily[f"d_{name}"] = daily["close"].ewm(span=span, adjust=False).mean()

    for window, name in [(SMA10_LEN, "sma10"), (SMA20_LEN, "sma20"), (SMA50_LEN, "sma50"),
                         (SMA100_LEN, "sma100"), (SMA200_LEN, "sma200")]:
        daily[f"d_{name}"] = daily["close"].rolling(window).mean()

    daily["d_vol10"] = daily["volume"].rolling(10).mean()
    daily["d_vol20"] = daily["volume"].rolling(20).mean()
    daily["d_vol30"] = daily["volume"].rolling(30).mean()
    daily["d_volume_ok"] = (daily["volume"] > daily["d_vol10"]) | (daily["volume"] > daily["d_vol20"]) | (daily["volume"] > daily["d_vol30"])

    # Daily TF2 conditions
    daily["d_short_term_stacked"] = (daily["d_ema10"] > daily["d_ema20"]) & (daily["d_ema20"] > daily["d_ema50"])
    daily["d_price_above_short"] = (daily["close"] > daily["d_ema10"]) & (daily["close"] > daily["d_ema20"]) & (daily["close"] > daily["d_ema50"])
    daily["d_short_term_ok"] = daily["d_short_term_stacked"] & daily["d_price_above_short"]
    daily["d_long_term_stacked"] = (daily["d_ema50"] > daily["d_ema100"]) & (daily["d_ema100"] > daily["d_ema200"])
    daily["d_ema50_rising"] = daily["d_ema50"] > daily["d_ema50"].shift(SLOPE_LEN)
    daily["d_ema100_rising"] = daily["d_ema100"] > daily["d_ema100"].shift(SLOPE_LEN)
    daily["d_price_above_long"] = (daily["close"] > daily["d_ema50"]) & (daily["close"] > daily["d_ema100"]) & (daily["close"] > daily["d_ema200"])
    daily["d_major_trend_ok"] = daily["d_long_term_stacked"] & daily["d_ema50_rising"] & daily["d_ema100_rising"] & daily["d_price_above_long"]
    daily["d_ema_uptrend"] = daily["d_ema9"] > daily["d_ema21"]
    daily["d_price_above_ema"] = daily["close"] > daily["d_ema21"]
    daily["d_tf2_short_ok"] = ~TF2_SHORT | daily["d_short_term_ok"]
    daily["d_tf2_long_ok"] = ~TF2_LONG | daily["d_major_trend_ok"]
    daily["d_tf2_ema_ok"] = daily["d_ema_uptrend"] & daily["d_price_above_ema"]
    daily["d_tf2_ok"] = daily["d_tf2_short_ok"] & daily["d_tf2_long_ok"] & daily["d_tf2_ema_ok"] & daily["d_volume_ok"]

    # Weekly resample
    daily_idx = daily.set_index("date")
    weekly = daily_idx.resample("W-FRI").agg({
        "open": "first", "high": "max", "low": "min", "close": "last", "volume": "sum"
    }).dropna().reset_index()
    weekly = weekly.sort_values("date").reset_index(drop=True)

    # Weekly EMAs / SMAs / ATR / Volume / Darvas
    for span, name in [(EMA1_LEN, "ema9"), (EMA2_LEN, "ema21"), (SMA10_LEN, "ema10"),
                       (SMA20_LEN, "ema20"), (SMA50_LEN, "ema50"), (SMA100_LEN, "ema100"), (SMA200_LEN, "ema200")]:
        weekly[f"w_{name}"] = weekly["close"].ewm(span=span, adjust=False).mean()

    for window, name in [(SMA10_LEN, "sma10"), (SMA20_LEN, "sma20"), (SMA50_LEN, "sma50"),
                         (SMA100_LEN, "sma100"), (SMA200_LEN, "sma200")]:
        weekly[f"w_{name}"] = weekly["close"].rolling(window).mean()

    prev_close = weekly["close"].shift(1)
    weekly["w_tr"] = pd.concat([
        weekly["high"] - weekly["low"],
        (weekly["high"] - prev_close).abs(),
        (weekly["low"] - prev_close).abs()
    ], axis=1).max(axis=1)
    weekly["w_atr"] = weekly["w_tr"].rolling(ATR_LEN).mean()

    weekly["w_vol10"] = weekly["volume"].rolling(10).mean()
    weekly["w_vol20"] = weekly["volume"].rolling(20).mean()
    weekly["w_vol30"] = weekly["volume"].rolling(30).mean()
    weekly["w_volume_ok"] = (weekly["volume"] > weekly["w_vol10"]) | (weekly["volume"] > weekly["w_vol20"]) | (weekly["volume"] > weekly["w_vol30"])

    body_top = weekly[["open", "close"]].max(axis=1)
    body_bottom = weekly[["open", "close"]].min(axis=1)
    weekly["w_darvas_top"] = body_top.rolling(GLOBAL_DARVAS1_LEN).max().shift(1)
    weekly["w_darvas_bottom"] = body_bottom.rolling(GLOBAL_DARVAS1_LEN).min().shift(1)
    weekly["w_darvas_ok"] = weekly["close"] > weekly["w_darvas_top"]

    # Weekly conditions
    weekly["w_ema_uptrend"] = weekly["w_ema9"] > weekly["w_ema21"]
    weekly["w_price_above_ema"] = weekly["close"] > weekly["w_ema21"]
    weekly["w_short_term_stacked"] = (weekly["w_ema10"] > weekly["w_ema20"]) & (weekly["w_ema20"] > weekly["w_ema50"])
    weekly["w_price_above_short"] = (weekly["close"] > weekly["w_ema10"]) & (weekly["close"] > weekly["w_ema20"]) & (weekly["close"] > weekly["w_ema50"])
    weekly["w_short_term_ok"] = weekly["w_short_term_stacked"] & weekly["w_price_above_short"]
    weekly["w_long_term_stacked"] = (weekly["w_ema50"] > weekly["w_ema100"]) & (weekly["w_ema100"] > weekly["w_ema200"])
    weekly["w_ema50_rising"] = weekly["w_ema50"] > weekly["w_ema50"].shift(SLOPE_LEN)
    weekly["w_ema100_rising"] = weekly["w_ema100"] > weekly["w_ema100"].shift(SLOPE_LEN)
    weekly["w_ema200_rising"] = weekly["w_ema200"] > weekly["w_ema200"].shift(SLOPE_LEN)
    weekly["w_core_rising"] = weekly["w_ema50_rising"] & weekly["w_ema100_rising"] & weekly["w_ema200_rising"]
    weekly["w_price_above_long"] = (weekly["close"] > weekly["w_ema50"]) & (weekly["close"] > weekly["w_ema100"]) & (weekly["close"] > weekly["w_ema200"])
    weekly["w_major_trend_ok"] = weekly["w_long_term_stacked"] & weekly["w_ema50_rising"] & weekly["w_ema100_rising"] & weekly["w_price_above_long"]
    weekly["w_full_bull"] = weekly["w_short_term_stacked"] & weekly["w_long_term_stacked"] & weekly["w_price_above_short"] & weekly["w_price_above_long"]

    weekly["w_base_core"] = weekly["w_ema_uptrend"] & weekly["w_price_above_ema"] & weekly["w_volume_ok"]
    if STRONG_ENTRY_CANDLE:
        cr = weekly["high"] - weekly["low"]
        br = (weekly["close"] - weekly["open"]).abs()
        bp = np.where(cr > 0, br / cr * 100, 0)
        cfh = np.where(cr > 0, (weekly["high"] - weekly["close"]) / cr * 100, 100)
        strong = (weekly["close"] > weekly["open"]) & (bp >= 40) & (cfh <= 25)
        weekly["w_base_core"] = weekly["w_base_core"] & strong

    weekly["w_tf1_short_ok"] = ~TF1_SHORT | weekly["w_short_term_ok"]
    weekly["w_tf1_long_ok"] = ~TF1_LONG | weekly["w_major_trend_ok"]
    weekly["w_tf1_ema_ok"] = weekly["w_ema_uptrend"] & weekly["w_price_above_ema"]
    weekly["w_tf1_ok"] = weekly["w_tf1_short_ok"] & weekly["w_tf1_long_ok"] & weekly["w_tf1_ema_ok"] & weekly["w_volume_ok"]
    weekly["w_global_darvas_ok"] = ~GLOBAL_DARVAS1_ON | weekly["w_darvas_ok"]

    # Crossunders for trend break
    weekly["w_ema_crossunder"] = (weekly["w_ema9"].shift(1) > weekly["w_ema21"].shift(1)) & (weekly["w_ema9"] <= weekly["w_ema21"])
    weekly["w_sma_crossunder"] = (weekly["w_sma50"].shift(1) > weekly["w_sma100"].shift(1)) & (weekly["w_sma50"] <= weekly["w_sma100"])

    # Map daily Friday values to weekly
    daily_last = daily.set_index("date")
    weekly = weekly.set_index("date")
    for col in ["d_tf2_ok", "d_ema_uptrend", "d_short_term_ok", "d_price_above_ema", "d_volume_ok",
                "d_ema9", "d_ema21", "d_ema50", "d_ema10", "d_ema20"]:
        weekly[f"m_{col}"] = daily_last[col].reindex(weekly.index, method="ffill")
    weekly = weekly.reset_index()

    # Full entry signal
    weekly["entry_signal"] = (
        weekly["w_base_core"] &
        weekly["w_tf1_ok"] &
        weekly["m_d_tf2_ok"] &
        weekly["w_global_darvas_ok"]
    )

    # Consolidation SL
    weekly["consolidation_sl"] = weekly["low"].rolling(5).min() - weekly["w_atr"] * SL_BUFFER_ATR

    return weekly


# =============================================================================
# BACKTEST ENGINE
# =============================================================================

@dataclass
class StaircasePosition:
    symbol: str
    level: str          # L1, L2, L3
    entry_date: str
    entry_price: float
    sl_price: float
    tp_price: float
    qty_pct: float
    exit_date: Optional[str] = None
    exit_price: Optional[float] = None
    exit_reason: Optional[str] = None
    partial_date: Optional[str] = None
    partial_price: Optional[float] = None
    partial_return_pct: Optional[float] = None
    final_return_pct: Optional[float] = None
    r_multiple: Optional[float] = None


def backtest_stock(symbol: str, weekly_df: pd.DataFrame) -> List[StaircasePosition]:
    """Run full staircase backtest on weekly data for one stock."""
    positions: List[StaircasePosition] = []

    # State
    pos_count = 0
    last_entry_bar = -999
    last_entry_bar_processed = -1

    entry1_price = entry1_sl = entry1_tp = None
    entry1_partial_taken = False
    entry1_bar = None

    entry2_price = entry2_sl = entry2_tp = None
    entry2_partial_taken = False
    entry2_bar = None

    entry3_price = entry3_sl = entry3_tp = None
    entry3_partial_taken = False
    entry3_bar = None

    for i, row in weekly_df.iterrows():
        if i < 50:
            continue  # Need enough history for indicators

        date = row["date"].strftime("%Y-%m-%d")
        close = row["close"]

        # Helper: current PnL % of each position
        def pnl_pct(entry_price):
            return (close - entry_price) / entry_price * 100 if entry_price else -999

        # --- 1. Trend break exit ---
        if pos_count > 0:
            if row["w_ema_crossunder"] or row["w_sma_crossunder"]:
                reason = "EMA_TrendBreak" if row["w_ema_crossunder"] else "SMA_TrendBreak"
                # Close all open positions
                for pos in positions:
                    if pos.exit_date is None:
                        pos.exit_date = date
                        pos.exit_price = round(close, 2)
                        pos.exit_reason = reason
                        # Calculate final return
                        if pos.partial_date:
                            # 50% at partial price, 50% at exit
                            pos.final_return_pct = round(
                                0.5 * (pos.partial_price - pos.entry_price) / pos.entry_price * 100 +
                                0.5 * (close - pos.entry_price) / pos.entry_price * 100, 2
                            )
                        else:
                            pos.final_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
                        risk = pos.entry_price - pos.sl_price
                        pos.r_multiple = round((close - pos.entry_price) / risk, 2) if risk and risk > 0 else 0
                pos_count = 0
                last_entry_bar = -999
                last_entry_bar_processed = -1
                entry1_price = entry1_sl = entry1_tp = None
                entry1_partial_taken = False
                entry1_bar = None
                entry2_price = entry2_sl = entry2_tp = None
                entry2_partial_taken = False
                entry2_bar = None
                entry3_price = entry3_sl = entry3_tp = None
                entry3_partial_taken = False
                entry3_bar = None
                continue

        # --- 2. Partial profit taking ---
        if pos_count >= 1 and entry1_tp and not entry1_partial_taken and close >= entry1_tp:
            for pos in positions:
                if pos.level == "L1" and pos.exit_date is None and pos.partial_date is None:
                    pos.partial_date = date
                    pos.partial_price = round(close, 2)
                    pos.partial_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
            entry1_partial_taken = True

        if pos_count >= 2 and entry2_tp and not entry2_partial_taken and close >= entry2_tp:
            for pos in positions:
                if pos.level == "L2" and pos.exit_date is None and pos.partial_date is None:
                    pos.partial_date = date
                    pos.partial_price = round(close, 2)
                    pos.partial_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
            entry2_partial_taken = True

        if pos_count >= 3 and entry3_tp and not entry3_partial_taken and close >= entry3_tp:
            for pos in positions:
                if pos.level == "L3" and pos.exit_date is None and pos.partial_date is None:
                    pos.partial_date = date
                    pos.partial_price = round(close, 2)
                    pos.partial_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
            entry3_partial_taken = True

        # --- 3. Stop loss hit ---
        composite_sl = None
        if pos_count >= 3 and entry3_sl:
            composite_sl = entry3_sl
        elif pos_count >= 2 and entry2_sl:
            composite_sl = entry2_sl
        elif pos_count >= 1 and entry1_sl:
            composite_sl = entry1_sl

        if pos_count > 0 and composite_sl and close < composite_sl:
            for pos in positions:
                if pos.exit_date is None:
                    pos.exit_date = date
                    pos.exit_price = round(close, 2)
                    pos.exit_reason = "SL_Hit"
                    if pos.partial_date:
                        pos.final_return_pct = round(
                            0.5 * (pos.partial_price - pos.entry_price) / pos.entry_price * 100 +
                            0.5 * (close - pos.entry_price) / pos.entry_price * 100, 2
                        )
                    else:
                        pos.final_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
                    risk = pos.entry_price - pos.sl_price
                    pos.r_multiple = round((close - pos.entry_price) / risk, 2) if risk and risk > 0 else 0
            pos_count = 0
            last_entry_bar = -999
            last_entry_bar_processed = -1
            entry1_price = entry1_sl = entry1_tp = None
            entry1_partial_taken = False
            entry1_bar = None
            entry2_price = entry2_sl = entry2_tp = None
            entry2_partial_taken = False
            entry2_bar = None
            entry3_price = entry3_sl = entry3_tp = None
            entry3_partial_taken = False
            entry3_bar = None
            continue

        # --- 4. Entries ---
        signal = row["entry_signal"]
        can_enter_new = (i - last_entry_bar) >= MIN_BARS_BETWEEN_ENTRIES
        can_take_new = i != last_entry_bar_processed

        if signal and can_take_new and pos_count == 0 and entry1_price is None:
            # L1 entry
            sl = row["consolidation_sl"]
            if pd.isna(sl):
                continue
            tp = close + (close - sl) * RISK_REWARD
            pos = StaircasePosition(
                symbol=symbol, level="L1", entry_date=date,
                entry_price=round(close, 2), sl_price=round(sl, 2),
                tp_price=round(tp, 2), qty_pct=POS_SIZE_L1
            )
            positions.append(pos)
            entry1_price = close
            entry1_sl = sl
            entry1_tp = tp
            entry1_partial_taken = False
            entry1_bar = i
            pos_count = 1
            last_entry_bar = i
            last_entry_bar_processed = i

        elif signal and can_take_new and pos_count == 1 and entry1_price is not None and entry1_bar is not None:
            # Check L2 add-on rules
            if i > entry1_bar and pnl_pct(entry1_price) >= MIN_PROFIT_FOR_NEXT_ENTRY:
                sl = row["consolidation_sl"]
                if pd.isna(sl):
                    continue
                tp = close + (close - sl) * RISK_REWARD
                pos = StaircasePosition(
                    symbol=symbol, level="L2", entry_date=date,
                    entry_price=round(close, 2), sl_price=round(sl, 2),
                    tp_price=round(tp, 2), qty_pct=POS_SIZE_L2
                )
                positions.append(pos)
                entry2_price = close
                entry2_sl = sl
                entry2_tp = tp
                entry2_partial_taken = False
                entry2_bar = i
                # Update L1 SL to consolidation SL
                entry1_sl = row["consolidation_sl"]
                for p in reversed(positions):
                    if p.level == "L1" and p.exit_date is None:
                        p.sl_price = round(row["consolidation_sl"], 2)
                        break
                pos_count = 2
                last_entry_bar = i
                last_entry_bar_processed = i

        elif signal and can_take_new and pos_count == 2 and entry2_price is not None and entry2_bar is not None:
            # Check L3 add-on rules
            if i > entry2_bar and pnl_pct(entry2_price) >= MIN_PROFIT_FOR_NEXT_ENTRY:
                sl = row["consolidation_sl"]
                if pd.isna(sl):
                    continue
                tp = close + (close - sl) * RISK_REWARD
                pos = StaircasePosition(
                    symbol=symbol, level="L3", entry_date=date,
                    entry_price=round(close, 2), sl_price=round(sl, 2),
                    tp_price=round(tp, 2), qty_pct=POS_SIZE_L3
                )
                positions.append(pos)
                entry3_price = close
                entry3_sl = sl
                entry3_tp = tp
                entry3_partial_taken = False
                entry3_bar = i
                # Update L1 and L2 SL to consolidation SL
                entry1_sl = row["consolidation_sl"]
                entry2_sl = row["consolidation_sl"]
                for p in reversed(positions):
                    if p.level in ("L1", "L2") and p.exit_date is None:
                        p.sl_price = round(row["consolidation_sl"], 2)
                pos_count = 3
                last_entry_bar = i
                last_entry_bar_processed = i

    # --- End of data: close any open positions ---
    if pos_count > 0:
        last_close = weekly_df.iloc[-1]["close"]
        last_date = weekly_df.iloc[-1]["date"].strftime("%Y-%m-%d")
        for pos in positions:
            if pos.exit_date is None:
                pos.exit_date = last_date
                pos.exit_price = round(last_close, 2)
                pos.exit_reason = "EndOfData"
                if pos.partial_date:
                    pos.final_return_pct = round(
                        0.5 * (pos.partial_price - pos.entry_price) / pos.entry_price * 100 +
                        0.5 * (last_close - pos.entry_price) / pos.entry_price * 100, 2
                    )
                else:
                    pos.final_return_pct = round((last_close - pos.entry_price) / pos.entry_price * 100, 2)
                risk = pos.entry_price - pos.sl_price
                pos.r_multiple = round((last_close - pos.entry_price) / risk, 2) if risk and risk > 0 else 0

    return positions


# =============================================================================
# EXCEL BUILDER
# =============================================================================

def style_header(ws, color):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def build_excel(all_positions: List[StaircasePosition]):
    wb = Workbook()
    wb.remove(wb.active)

    # Prepare dataframes
    df_pos = pd.DataFrame([asdict(p) for p in all_positions])

    # Sheet 1: Trade_Summary (all round trips)
    ws1 = wb.create_sheet("Trade_Summary")
    for r in dataframe_to_rows(df_pos, index=False, header=True):
        ws1.append(r)
    style_header(ws1, "2F5496")
    auto_width(ws1)

    # Sheet 2: Entry_Log
    entries = df_pos[["symbol", "level", "entry_date", "entry_price", "sl_price", "tp_price", "qty_pct"]].copy()
    ws2 = wb.create_sheet("Entry_Log")
    for r in dataframe_to_rows(entries, index=False, header=True):
        ws2.append(r)
    style_header(ws2, "375623")
    auto_width(ws2)

    # Sheet 3: Partial_Log
    partials = df_pos[df_pos["partial_date"].notna()][[
        "symbol", "level", "entry_date", "entry_price", "partial_date", "partial_price", "partial_return_pct"
    ]].copy()
    ws3 = wb.create_sheet("Partial_Log")
    for r in dataframe_to_rows(partials, index=False, header=True):
        ws3.append(r)
    style_header(ws3, "C55A11")
    auto_width(ws3)

    # Sheet 4: Exit_Events
    exits = df_pos[df_pos["exit_date"].notna()][[
        "symbol", "level", "entry_date", "entry_price", "exit_date", "exit_price",
        "exit_reason", "final_return_pct", "r_multiple"
    ]].copy()
    ws4 = wb.create_sheet("Exit_Events")
    for r in dataframe_to_rows(exits, index=False, header=True):
        ws4.append(r)
    style_header(ws4, "7030A0")
    auto_width(ws4)

    # Sheet 5: Stock_Summary
    summary_rows = []
    for sym, g in df_pos.groupby("symbol"):
        total_trades = len(g)
        l1_count = len(g[g["level"] == "L1"])
        l2_count = len(g[g["level"] == "L2"])
        l3_count = len(g[g["level"] == "L3"])
        partials_taken = g["partial_date"].notna().sum()
        wins = (g["final_return_pct"] > 0).sum()
        losses = (g["final_return_pct"] < 0).sum()
        win_rate = round(wins / total_trades * 100, 1) if total_trades > 0 else 0
        avg_return = round(g["final_return_pct"].mean(), 2)
        total_r = round(g["r_multiple"].sum(), 2)
        avg_r = round(g["r_multiple"].mean(), 2)
        max_r = round(g["r_multiple"].max(), 2)
        min_r = round(g["r_multiple"].min(), 2)
        summary_rows.append({
            "symbol": sym,
            "total_trades": total_trades,
            "L1": l1_count,
            "L2": l2_count,
            "L3": l3_count,
            "partials": int(partials_taken),
            "wins": int(wins),
            "losses": int(losses),
            "win_rate_pct": win_rate,
            "avg_return_pct": avg_return,
            "total_r": total_r,
            "avg_r": avg_r,
            "max_r": max_r,
            "min_r": min_r,
        })

    df_summary = pd.DataFrame(summary_rows)
    ws5 = wb.create_sheet("Stock_Summary")
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws5.append(r)
    style_header(ws5, "1F4E78")
    auto_width(ws5)

    wb.save(OUTPUT_EXCEL)
    print(f"\n[OK] Excel saved: {OUTPUT_EXCEL}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 70)
    print("STAIRCASE STRATEGY WEEKLY BACKTEST ENGINE")
    print("=" * 70)

    all_positions: List[StaircasePosition] = []

    for symbol in SYMBOLS:
        print(f"\nBacktesting {symbol}...")
        try:
            daily_df = load_daily_data(symbol)
            weekly_df = compute_indicators(daily_df)
            positions = backtest_stock(symbol, weekly_df)
            all_positions.extend(positions)
            print(f"  -> {len(positions)} positions generated")
        except Exception as e:
            print(f"  [ERROR] {symbol}: {e}")

    print(f"\n{'='*70}")
    print(f"TOTAL POSITIONS: {len(all_positions)}")
    print(f"{'='*70}")

    build_excel(all_positions)


if __name__ == "__main__":
    from dataclasses import asdict
    main()
