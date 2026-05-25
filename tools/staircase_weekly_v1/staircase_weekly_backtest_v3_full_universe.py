#!/usr/bin/env python3
"""
Staircase Strategy Weekly Backtest Engine V3 — Full Universe
=============================================================
Runs the V3 strategy (valid days REMOVED, point-in-time TF checks only)
on the full 734-symbol Dhan dataset.

Data source: data/nse_index_history/run_20260521_194407/history/
"""
from __future__ import annotations

import json
import warnings
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")

# =============================================================================
# PATHS & CONFIG
# =============================================================================
ROOT = Path(__file__).resolve().parents[1]
HISTORY_DIR = ROOT / "data" / "nse_index_history" / "run_20260521_194407" / "history"
SYMBOLS_CSV = ROOT / "data" / "nse_index_history" / "run_20260521_194407" / "symbols.csv"
OUTPUT_DIR = ROOT / "analysis" / "staircase_dhan"
OUTPUT_EXCEL = OUTPUT_DIR / "staircase_strategy_weekly_backtest_v3_full_universe.xlsx"
OUTPUT_JSON = OUTPUT_DIR / "staircase_strategy_weekly_backtest_v3_full_universe.json"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# --- USER TV SETTINGS (from input panel screenshot) ---
POS_SIZE_L1 = 50.0
POS_SIZE_L2 = 30.0
POS_SIZE_L3 = 20.0
MIN_BARS_BETWEEN_ENTRIES = 1
VOLUME_FILTER_ON = True
VOLUME_REQUIREMENT = "Above Any (10/20/30)"
STRONG_ENTRY_CANDLE = False
MIN_ENTRY_BODY_RANGE_PCT = 60.0
MAX_ENTRY_CLOSE_FROM_HIGH_PCT = 20.0
SL_TYPE = "Consolidation Low"
SL_BUFFER_ATR = 0.2
RISK_REWARD_L1 = 2.0
RISK_REWARD_L2 = 2.0
RISK_REWARD_L3 = 2.0
PARTIAL_PCT = 50.0
MIN_PROFIT_FOR_NEXT_ENTRY = 0.0

# --- TIMEFRAME FILTER SETTINGS ---
# TF1 = Weekly (chart timeframe)
TF1_ON = True
TF1_SHORT = True
TF1_LONG = False
TF1_DARVAS = False
TF1_DARVAS_UPPER_LEN = 15
TF1_DARVAS_LOWER_LEN = 15

# TF2 = Daily
TF2_ON = True
TF2_SHORT = True
TF2_LONG = True
TF2_DARVAS = False
TF2_DARVAS_UPPER_LEN = 15
TF2_DARVAS_LOWER_LEN = 15

# Global Darvas
GLOBAL_DARVAS1_ON = True
GLOBAL_DARVAS1_LEN = 40

# --- INDICATOR SETTINGS ---
EMA1_LEN, EMA2_LEN = 9, 21
SMA10_LEN, SMA20_LEN, SMA50_LEN = 10, 20, 50
SMA100_LEN, SMA200_LEN = 100, 200
SLOPE_LEN = 5
ATR_LEN = 14

MIN_DAILY_ROWS = 250  # Need enough data for 200 EMA + warmup

# =============================================================================
# DATA LOADING
# =============================================================================

def get_symbols_with_history() -> List[str]:
    """Load symbols that have historical data available."""
    df = pd.read_csv(SYMBOLS_CSV)
    symbols = df[df["has_history"] == True]["symbol"].tolist()
    return sorted(symbols)


def load_dhan_history(symbol: str) -> Optional[pd.DataFrame]:
    """Load daily OHLCV from Dhan history CSV."""
    path = HISTORY_DIR / f"{symbol}.csv"
    if not path.exists():
        return None
    df = pd.read_csv(path)
    # Rename columns to lowercase for compatibility
    df.columns = [c.lower() for c in df.columns]
    df = df.rename(columns={"date": "date"})
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    # Ensure required columns exist
    required = ["open", "high", "low", "close", "volume"]
    for col in required:
        if col not in df.columns:
            return None
    df["volume"] = df["volume"].fillna(0)
    return df[["date", "open", "high", "low", "close", "volume"]]


# =============================================================================
# INDICATORS
# =============================================================================

def calc_volume_ok(df: pd.DataFrame) -> pd.Series:
    """Volume filter matching Pine calcVolumeOK()."""
    if not VOLUME_FILTER_ON:
        return pd.Series(True, index=df.index)
    vol10 = df["volume"].rolling(10, min_periods=10).mean()
    vol20 = df["volume"].rolling(20, min_periods=20).mean()
    vol30 = df["volume"].rolling(30, min_periods=30).mean()
    above10 = df["volume"] > vol10
    above20 = df["volume"] > vol20
    above30 = df["volume"] > vol30
    if VOLUME_REQUIREMENT == "Above All (10/20/30)":
        return above10 & above20 & above30
    if VOLUME_REQUIREMENT == "Above Any (10/20/30)":
        return above10 | above20 | above30
    return above10


def add_indicators(df: pd.DataFrame, prefix: str = "") -> pd.DataFrame:
    """Add EMAs, SMAs, ATR, volume, trend conditions to a dataframe."""
    out = df.copy()
    # EMAs
    for span, name in [(EMA1_LEN, "ema9"), (EMA2_LEN, "ema21"),
                       (SMA10_LEN, "ema10"), (SMA20_LEN, "ema20"),
                       (SMA50_LEN, "ema50"), (SMA100_LEN, "ema100"), (SMA200_LEN, "ema200")]:
        out[f"{prefix}{name}"] = out["close"].ewm(span=span, adjust=False).mean()
    # SMAs (for SL and crossunder only)
    for window, name in [(SMA10_LEN, "sma10"), (SMA20_LEN, "sma20"), (SMA50_LEN, "sma50"),
                         (SMA100_LEN, "sma100"), (SMA200_LEN, "sma200")]:
        out[f"{prefix}{name}"] = out["close"].rolling(window, min_periods=window).mean()
    # ATR
    prev_close = out["close"].shift(1)
    tr = pd.concat([
        out["high"] - out["low"],
        (out["high"] - prev_close).abs(),
        (out["low"] - prev_close).abs()
    ], axis=1).max(axis=1)
    out[f"{prefix}atr"] = tr.rolling(ATR_LEN, min_periods=ATR_LEN).mean()
    # Volume
    out[f"{prefix}volume_ok"] = calc_volume_ok(out)
    # Trend conditions (using EMAs, matching Pine calcTFData)
    out[f"{prefix}short_term_stacked"] = (out[f"{prefix}ema10"] > out[f"{prefix}ema20"]) & (out[f"{prefix}ema20"] > out[f"{prefix}ema50"])
    out[f"{prefix}price_above_short"] = (out["close"] > out[f"{prefix}ema10"]) & (out["close"] > out[f"{prefix}ema20"]) & (out["close"] > out[f"{prefix}ema50"])
    out[f"{prefix}short_term_ok"] = out[f"{prefix}short_term_stacked"] & out[f"{prefix}price_above_short"]
    out[f"{prefix}long_term_stacked"] = (out[f"{prefix}ema50"] > out[f"{prefix}ema100"]) & (out[f"{prefix}ema100"] > out[f"{prefix}ema200"])
    out[f"{prefix}ema50_rising"] = out[f"{prefix}ema50"] > out[f"{prefix}ema50"].shift(SLOPE_LEN)
    out[f"{prefix}ema100_rising"] = out[f"{prefix}ema100"] > out[f"{prefix}ema100"].shift(SLOPE_LEN)
    out[f"{prefix}price_above_long"] = (out["close"] > out[f"{prefix}ema50"]) & (out["close"] > out[f"{prefix}ema100"]) & (out["close"] > out[f"{prefix}ema200"])
    out[f"{prefix}major_trend_ok"] = out[f"{prefix}long_term_stacked"] & out[f"{prefix}ema50_rising"] & out[f"{prefix}ema100_rising"] & out[f"{prefix}price_above_long"]
    out[f"{prefix}ema_uptrend"] = out[f"{prefix}ema9"] > out[f"{prefix}ema21"]
    out[f"{prefix}price_above_ema"] = out["close"] > out[f"{prefix}ema21"]
    # Darvas
    body_top = out[["open", "close"]].max(axis=1)
    body_bottom = out[["open", "close"]].min(axis=1)
    out[f"{prefix}darvas_top"] = body_top.rolling(GLOBAL_DARVAS1_LEN, min_periods=GLOBAL_DARVAS1_LEN).max().shift(1)
    out[f"{prefix}darvas_bottom"] = body_bottom.rolling(GLOBAL_DARVAS1_LEN, min_periods=GLOBAL_DARVAS1_LEN).min().shift(1)
    out[f"{prefix}darvas_ok"] = out["close"] > out[f"{prefix}darvas_top"]
    # Crossunders
    out[f"{prefix}ema_crossunder"] = (out[f"{prefix}ema9"].shift(1) > out[f"{prefix}ema21"].shift(1)) & (out[f"{prefix}ema9"] < out[f"{prefix}ema21"])
    out[f"{prefix}sma_crossunder"] = (out[f"{prefix}sma50"].shift(1) > out[f"{prefix}sma100"].shift(1)) & (out[f"{prefix}sma50"] < out[f"{prefix}sma100"])
    # Strong candle
    if STRONG_ENTRY_CANDLE:
        cr = out["high"] - out["low"]
        br = (out["close"] - out["open"]).abs()
        bp = np.where(cr > 0, br / cr * 100, 0)
        cfh = np.where(cr > 0, (out["high"] - out["close"]) / cr * 100, 100)
        out[f"{prefix}strong_candle"] = (out["close"] > out["open"]) & (bp >= MIN_ENTRY_BODY_RANGE_PCT) & (cfh <= MAX_ENTRY_CLOSE_FROM_HIGH_PCT)
    else:
        out[f"{prefix}strong_candle"] = True
    return out


def compute_indicators(daily_df: pd.DataFrame) -> pd.DataFrame:
    """Compute weekly + daily indicators and merge into weekly dataframe."""
    daily = add_indicators(daily_df.copy(), "d_")

    # Daily TF2 signal (point-in-time, no valid days)
    daily["d_tf2_short_ok"] = ~TF2_SHORT | daily["d_short_term_ok"]
    daily["d_tf2_long_ok"] = ~TF2_LONG | daily["d_major_trend_ok"]
    daily["d_tf2_ema_ok"] = daily["d_ema_uptrend"] & daily["d_price_above_ema"]
    daily["d_tf2_ok"] = daily["d_tf2_short_ok"] & daily["d_tf2_long_ok"] & daily["d_tf2_ema_ok"] & daily["d_volume_ok"]
    if TF2_DARVAS:
        daily["d_tf2_ok"] = daily["d_tf2_ok"] & daily["d_darvas_ok"]

    # Weekly resample
    daily_idx = daily.set_index("date")
    weekly = daily_idx.resample("W-FRI").agg({
        "open": "first", "high": "max", "low": "min", "close": "last", "volume": "sum"
    }).dropna().reset_index()
    weekly = weekly.sort_values("date").reset_index(drop=True)

    # Weekly indicators
    weekly = add_indicators(weekly, "w_")

    # Weekly TF1 signal (point-in-time, no valid days)
    weekly["w_tf1_short_ok"] = ~TF1_SHORT | weekly["w_short_term_ok"]
    weekly["w_tf1_long_ok"] = ~TF1_LONG | weekly["w_major_trend_ok"]
    weekly["w_tf1_ema_ok"] = weekly["w_ema_uptrend"] & weekly["w_price_above_ema"]
    weekly["w_tf1_ok"] = weekly["w_tf1_short_ok"] & weekly["w_tf1_long_ok"] & weekly["w_tf1_ema_ok"] & weekly["w_volume_ok"]
    if TF1_DARVAS:
        weekly["w_tf1_ok"] = weekly["w_tf1_ok"] & weekly["w_darvas_ok"]

    # Chart base entry core
    weekly["w_base_core"] = weekly["w_ema_uptrend"] & weekly["w_price_above_ema"] & weekly["w_volume_ok"] & weekly["w_strong_candle"]

    # Global Darvas
    weekly["w_global_darvas_ok"] = ~GLOBAL_DARVAS1_ON | weekly["w_darvas_ok"]

    # Map daily Friday values to weekly
    daily_last = daily.set_index("date")
    weekly = weekly.set_index("date")
    for col in ["d_tf2_ok"]:
        weekly[f"m_{col}"] = daily_last[col].reindex(weekly.index, method="ffill")
    weekly = weekly.reset_index()

    # Full entry signal
    weekly["entry_signal"] = (
        weekly["w_base_core"] &
        weekly["w_tf1_ok"] &
        weekly["m_d_tf2_ok"] &
        weekly["w_global_darvas_ok"]
    )

    # SL calculation
    def calc_sl(row):
        if SL_TYPE == "Consolidation Low":
            return row["low"].rolling(5, min_periods=5).min() - row["w_atr"] * SL_BUFFER_ATR
        elif SL_TYPE == "SMA 50":
            return row["w_sma50"] - row["w_atr"] * SL_BUFFER_ATR
        elif SL_TYPE == "SMA 200":
            return row["w_sma200"] - row["w_atr"] * SL_BUFFER_ATR
        return row["low"].rolling(5, min_periods=5).min() - row["w_atr"] * SL_BUFFER_ATR

    weekly["consolidation_sl"] = calc_sl(weekly)

    return weekly


# =============================================================================
# BACKTEST ENGINE
# =============================================================================

@dataclass
class StaircasePosition:
    symbol: str
    level: str
    entry_date: str
    entry_price: float
    sl_price: float
    tp_price: float
    qty_pct: float
    exit_date: Optional[str] = None
    exit_price: Optional[float] = None
    exit_reason: Optional[str] = None
    partial_date: Optional[str] = None
    partial_price: Optional[float] = None
    partial_return_pct: Optional[float] = None
    final_return_pct: Optional[float] = None
    r_multiple: Optional[float] = None


def backtest_stock(symbol: str, weekly_df: pd.DataFrame) -> List[StaircasePosition]:
    """Run full staircase backtest on weekly data for one stock."""
    positions: List[StaircasePosition] = []

    pos_count = 0
    last_entry_bar = -999
    last_entry_bar_processed = -1

    entry1_price = entry1_sl = entry1_tp = None
    entry1_partial_taken = False
    entry1_bar = None

    entry2_price = entry2_sl = entry2_tp = None
    entry2_partial_taken = False
    entry2_bar = None

    entry3_price = entry3_sl = entry3_tp = None
    entry3_partial_taken = False
    entry3_bar = None

    partial_weight = PARTIAL_PCT / 100.0
    remainder_weight = 1.0 - partial_weight

    for i, row in weekly_df.iterrows():
        if i < 50:
            continue

        date = row["date"].strftime("%Y-%m-%d")
        close = row["close"]

        def pnl_pct(entry_price):
            return (close - entry_price) / entry_price * 100 if entry_price else -999

        # --- 1. Trend break exit ---
        if pos_count > 0:
            if row["w_ema_crossunder"] or row["w_sma_crossunder"]:
                reason = "EMA_TrendBreak" if row["w_ema_crossunder"] else "SMA_TrendBreak"
                for pos in positions:
                    if pos.exit_date is None:
                        pos.exit_date = date
                        pos.exit_price = round(close, 2)
                        pos.exit_reason = reason
                        if pos.partial_date:
                            pos.final_return_pct = round(
                                partial_weight * (pos.partial_price - pos.entry_price) / pos.entry_price * 100 +
                                remainder_weight * (close - pos.entry_price) / pos.entry_price * 100, 2
                            )
                        else:
                            pos.final_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
                        risk = pos.entry_price - pos.sl_price
                        pos.r_multiple = round((close - pos.entry_price) / risk, 2) if risk and risk > 0 else 0
                pos_count = 0
                last_entry_bar = -999
                last_entry_bar_processed = -1
                entry1_price = entry1_sl = entry1_tp = None
                entry1_partial_taken = False
                entry1_bar = None
                entry2_price = entry2_sl = entry2_tp = None
                entry2_partial_taken = False
                entry2_bar = None
                entry3_price = entry3_sl = entry3_tp = None
                entry3_partial_taken = False
                entry3_bar = None
                continue

        # --- 2. Partial profit taking ---
        if pos_count >= 1 and entry1_tp and not entry1_partial_taken and close >= entry1_tp:
            for pos in positions:
                if pos.level == "L1" and pos.exit_date is None and pos.partial_date is None:
                    pos.partial_date = date
                    pos.partial_price = round(close, 2)
                    pos.partial_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
            entry1_partial_taken = True

        if pos_count >= 2 and entry2_tp and not entry2_partial_taken and close >= entry2_tp:
            for pos in positions:
                if pos.level == "L2" and pos.exit_date is None and pos.partial_date is None:
                    pos.partial_date = date
                    pos.partial_price = round(close, 2)
                    pos.partial_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
            entry2_partial_taken = True

        if pos_count >= 3 and entry3_tp and not entry3_partial_taken and close >= entry3_tp:
            for pos in positions:
                if pos.level == "L3" and pos.exit_date is None and pos.partial_date is None:
                    pos.partial_date = date
                    pos.partial_price = round(close, 2)
                    pos.partial_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
            entry3_partial_taken = True

        # --- 3. Stop loss hit ---
        composite_sl = None
        if pos_count >= 3 and entry3_sl:
            composite_sl = entry3_sl
        elif pos_count >= 2 and entry2_sl:
            composite_sl = entry2_sl
        elif pos_count >= 1 and entry1_sl:
            composite_sl = entry1_sl

        if pos_count > 0 and composite_sl and close < composite_sl:
            for pos in positions:
                if pos.exit_date is None:
                    pos.exit_date = date
                    pos.exit_price = round(close, 2)
                    pos.exit_reason = "SL_Hit"
                    if pos.partial_date:
                        pos.final_return_pct = round(
                            partial_weight * (pos.partial_price - pos.entry_price) / pos.entry_price * 100 +
                            remainder_weight * (close - pos.entry_price) / pos.entry_price * 100, 2
                        )
                    else:
                        pos.final_return_pct = round((close - pos.entry_price) / pos.entry_price * 100, 2)
                    risk = pos.entry_price - pos.sl_price
                    pos.r_multiple = round((close - pos.entry_price) / risk, 2) if risk and risk > 0 else 0
            pos_count = 0
            last_entry_bar = -999
            last_entry_bar_processed = -1
            entry1_price = entry1_sl = entry1_tp = None
            entry1_partial_taken = False
            entry1_bar = None
            entry2_price = entry2_sl = entry2_tp = None
            entry2_partial_taken = False
            entry2_bar = None
            entry3_price = entry3_sl = entry3_tp = None
            entry3_partial_taken = False
            entry3_bar = None
            continue

        # --- 4. Entries ---
        signal = row["entry_signal"]
        can_enter_new = (i - last_entry_bar) >= MIN_BARS_BETWEEN_ENTRIES
        can_take_new = i != last_entry_bar_processed

        if signal and can_enter_new and can_take_new and pos_count == 0 and entry1_price is None:
            # L1 entry
            sl = row["consolidation_sl"]
            if pd.isna(sl):
                continue
            tp = close + (close - sl) * RISK_REWARD_L1
            pos = StaircasePosition(
                symbol=symbol, level="L1", entry_date=date,
                entry_price=round(close, 2), sl_price=round(sl, 2),
                tp_price=round(tp, 2), qty_pct=POS_SIZE_L1
            )
            positions.append(pos)
            entry1_price = close
            entry1_sl = sl
            entry1_tp = tp
            entry1_partial_taken = False
            entry1_bar = i
            pos_count = 1
            last_entry_bar = i
            last_entry_bar_processed = i

        elif signal and can_enter_new and can_take_new and pos_count == 1 and entry1_price is not None and entry1_bar is not None:
            # L2 add-on
            if i > entry1_bar and pnl_pct(entry1_price) >= MIN_PROFIT_FOR_NEXT_ENTRY:
                sl = row["consolidation_sl"]
                if pd.isna(sl):
                    continue
                tp = close + (close - sl) * RISK_REWARD_L2
                pos = StaircasePosition(
                    symbol=symbol, level="L2", entry_date=date,
                    entry_price=round(close, 2), sl_price=round(sl, 2),
                    tp_price=round(tp, 2), qty_pct=POS_SIZE_L2
                )
                positions.append(pos)
                entry2_price = close
                entry2_sl = sl
                entry2_tp = tp
                entry2_partial_taken = False
                entry2_bar = i
                # Update L1 SL to consolidation SL
                entry1_sl = row["consolidation_sl"]
                for p in reversed(positions):
                    if p.level == "L1" and p.exit_date is None:
                        p.sl_price = round(row["consolidation_sl"], 2)
                        break
                pos_count = 2
                last_entry_bar = i
                last_entry_bar_processed = i

        elif signal and can_enter_new and can_take_new and pos_count == 2 and entry2_price is not None and entry2_bar is not None:
            # L3 add-on
            if i > entry2_bar and pnl_pct(entry2_price) >= MIN_PROFIT_FOR_NEXT_ENTRY:
                sl = row["consolidation_sl"]
                if pd.isna(sl):
                    continue
                tp = close + (close - sl) * RISK_REWARD_L3
                pos = StaircasePosition(
                    symbol=symbol, level="L3", entry_date=date,
                    entry_price=round(close, 2), sl_price=round(sl, 2),
                    tp_price=round(tp, 2), qty_pct=POS_SIZE_L3
                )
                positions.append(pos)
                entry3_price = close
                entry3_sl = sl
                entry3_tp = tp
                entry3_partial_taken = False
                entry3_bar = i
                # Update L1 and L2 SL to consolidation SL
                entry1_sl = row["consolidation_sl"]
                entry2_sl = row["consolidation_sl"]
                for p in reversed(positions):
                    if p.level in ("L1", "L2") and p.exit_date is None:
                        p.sl_price = round(row["consolidation_sl"], 2)
                pos_count = 3
                last_entry_bar = i
                last_entry_bar_processed = i

    # --- End of data: close any open positions ---
    if pos_count > 0:
        last_close = weekly_df.iloc[-1]["close"]
        last_date = weekly_df.iloc[-1]["date"].strftime("%Y-%m-%d")
        for pos in positions:
            if pos.exit_date is None:
                pos.exit_date = last_date
                pos.exit_price = round(last_close, 2)
                pos.exit_reason = "EndOfData"
                if pos.partial_date:
                    pos.final_return_pct = round(
                        partial_weight * (pos.partial_price - pos.entry_price) / pos.entry_price * 100 +
                        remainder_weight * (last_close - pos.entry_price) / pos.entry_price * 100, 2
                    )
                else:
                    pos.final_return_pct = round((last_close - pos.entry_price) / pos.entry_price * 100, 2)
                risk = pos.entry_price - pos.sl_price
                pos.r_multiple = round((last_close - pos.entry_price) / risk, 2) if risk and risk > 0 else 0

    return positions


# =============================================================================
# EXCEL BUILDER
# =============================================================================

def style_header(ws, color):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def build_excel(all_positions: List[StaircasePosition]):
    wb = Workbook()
    wb.remove(wb.active)

    df_pos = pd.DataFrame([asdict(p) for p in all_positions])

    # Sheet 1: Trade_Summary
    ws1 = wb.create_sheet("Trade_Summary")
    for r in dataframe_to_rows(df_pos, index=False, header=True):
        ws1.append(r)
    style_header(ws1, "2F5496")
    auto_width(ws1)

    # Sheet 2: Entry_Log
    entries = df_pos[["symbol", "level", "entry_date", "entry_price", "sl_price", "tp_price", "qty_pct"]].copy()
    ws2 = wb.create_sheet("Entry_Log")
    for r in dataframe_to_rows(entries, index=False, header=True):
        ws2.append(r)
    style_header(ws2, "375623")
    auto_width(ws2)

    # Sheet 3: Partial_Log
    partials = df_pos[df_pos["partial_date"].notna()][[
        "symbol", "level", "entry_date", "entry_price", "partial_date", "partial_price", "partial_return_pct"
    ]].copy()
    ws3 = wb.create_sheet("Partial_Log")
    for r in dataframe_to_rows(partials, index=False, header=True):
        ws3.append(r)
    style_header(ws3, "C55A11")
    auto_width(ws3)

    # Sheet 4: Exit_Events
    exits = df_pos[df_pos["exit_date"].notna()][[
        "symbol", "level", "entry_date", "entry_price", "exit_date", "exit_price",
        "exit_reason", "final_return_pct", "r_multiple"
    ]].copy()
    ws4 = wb.create_sheet("Exit_Events")
    for r in dataframe_to_rows(exits, index=False, header=True):
        ws4.append(r)
    style_header(ws4, "7030A0")
    auto_width(ws4)

    # Sheet 5: Stock_Summary
    summary_rows = []
    for sym, g in df_pos.groupby("symbol"):
        total_trades = len(g)
        l1_count = len(g[g["level"] == "L1"])
        l2_count = len(g[g["level"] == "L2"])
        l3_count = len(g[g["level"] == "L3"])
        partials_taken = g["partial_date"].notna().sum()
        wins = (g["final_return_pct"] > 0).sum()
        losses = (g["final_return_pct"] < 0).sum()
        win_rate = round(wins / total_trades * 100, 1) if total_trades > 0 else 0
        avg_return = round(g["final_return_pct"].mean(), 2)
        total_r = round(g["r_multiple"].sum(), 2)
        avg_r = round(g["r_multiple"].mean(), 2)
        max_r = round(g["r_multiple"].max(), 2)
        min_r = round(g["r_multiple"].min(), 2)
        summary_rows.append({
            "symbol": sym,
            "total_trades": total_trades,
            "L1": l1_count,
            "L2": l2_count,
            "L3": l3_count,
            "partials": int(partials_taken),
            "wins": int(wins),
            "losses": int(losses),
            "win_rate_pct": win_rate,
            "avg_return_pct": avg_return,
            "total_r": total_r,
            "avg_r": avg_r,
            "max_r": max_r,
            "min_r": min_r,
        })

    df_summary = pd.DataFrame(summary_rows)
    ws5 = wb.create_sheet("Stock_Summary")
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws5.append(r)
    style_header(ws5, "1F4E78")
    auto_width(ws5)

    # Sheet 6: Universe_Overview
    overview = {
        "total_symbols_processed": len(df_summary),
        "total_positions": len(df_pos),
        "total_L1": int(df_summary["L1"].sum()),
        "total_L2": int(df_summary["L2"].sum()),
        "total_L3": int(df_summary["L3"].sum()),
        "total_partials": int(df_summary["partials"].sum()),
        "total_wins": int(df_summary["wins"].sum()),
        "total_losses": int(df_summary["losses"].sum()),
        "overall_win_rate": round(df_summary["wins"].sum() / len(df_pos) * 100, 1) if len(df_pos) > 0 else 0,
        "avg_return_all": round(df_pos["final_return_pct"].mean(), 2),
        "total_r_all": round(df_pos["r_multiple"].sum(), 2),
        "avg_r_all": round(df_pos["r_multiple"].mean(), 2),
        "symbols_with_trades": int((df_summary["total_trades"] > 0).sum()),
        "symbols_without_trades": int((df_summary["total_trades"] == 0).sum()),
    }
    ws6 = wb.create_sheet("Universe_Overview")
    ws6.append(["Metric", "Value"])
    for k, v in overview.items():
        ws6.append([k, v])
    style_header(ws6, "2E75B6")
    auto_width(ws6)

    wb.save(OUTPUT_EXCEL)
    print(f"\n[OK] Excel saved: {OUTPUT_EXCEL}")

    # Also save JSON summary
    with open(OUTPUT_JSON, "w") as f:
        json.dump({
            "overview": overview,
            "stock_summary": df_summary.to_dict(orient="records"),
        }, f, indent=2, default=str)
    print(f"[OK] JSON saved: {OUTPUT_JSON}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    symbols = get_symbols_with_history()
    print("=" * 70)
    print("STAIRCASE STRATEGY WEEKLY BACKTEST ENGINE V3 — FULL UNIVERSE")
    print("=" * 70)
    print(f"Data Source: Dhan history ({HISTORY_DIR})")
    print(f"Symbols with history: {len(symbols)}")
    print(f"Scaling: {POS_SIZE_L1:.0f}/{POS_SIZE_L2:.0f}/{POS_SIZE_L3:.0f}")
    print(f"Min Bars Between Entries: {MIN_BARS_BETWEEN_ENTRIES}")
    print(f"Volume Filter: {VOLUME_REQUIREMENT} ({'ON' if VOLUME_FILTER_ON else 'OFF'})")
    print(f"Strong Candle: {'ON' if STRONG_ENTRY_CANDLE else 'OFF'}")
    print(f"SL Type: {SL_TYPE}")
    print(f"TF1: Weekly, Short={'ON' if TF1_SHORT else 'OFF'}, Long={'ON' if TF1_LONG else 'OFF'}, Darvas={'ON' if TF1_DARVAS else 'OFF'}")
    print(f"TF2: Daily, Short={'ON' if TF2_SHORT else 'OFF'}, Long={'ON' if TF2_LONG else 'OFF'}, Darvas={'ON' if TF2_DARVAS else 'OFF'}")
    print(f"Global Darvas 1: {'ON' if GLOBAL_DARVAS1_ON else 'OFF'}, len={GLOBAL_DARVAS1_LEN}")
    print(f"Valid Days: REMOVED (point-in-time only)")
    print("=" * 70)

    all_positions: List[StaircasePosition] = []
    skipped = []
    errors = []

    for idx, symbol in enumerate(symbols, 1):
        try:
            daily_df = load_dhan_history(symbol)
            if daily_df is None:
                skipped.append((symbol, "no_data"))
                continue
            if len(daily_df) < MIN_DAILY_ROWS:
                skipped.append((symbol, f"insufficient_data ({len(daily_df)} rows)"))
                continue

            weekly_df = compute_indicators(daily_df)
            positions = backtest_stock(symbol, weekly_df)
            all_positions.extend(positions)

            if idx % 50 == 0 or idx == len(symbols):
                print(f"  ... {idx}/{len(symbols)} ({symbol}: {len(positions)} positions, total so far: {len(all_positions)})")
        except Exception as e:
            errors.append((symbol, str(e)))
            if idx % 50 == 0:
                print(f"  ... {idx}/{len(symbols)} (error on {symbol}: {e})")

    print(f"\n{'='*70}")
    print(f"TOTAL POSITIONS: {len(all_positions)}")
    print(f"Symbols skipped: {len(skipped)}")
    print(f"Symbols with errors: {len(errors)}")
    if errors:
        for sym, err in errors[:10]:
            print(f"  - {sym}: {err}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more errors")
    print(f"{'='*70}")

    build_excel(all_positions)


if __name__ == "__main__":
    main()
